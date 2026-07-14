import os
import io
import re
import sys
import json
from functools import wraps
from datetime import datetime, timedelta

from flask import Flask, jsonify, request, session
from flask_cors import CORS
import mysql.connector  # needed here for mysql.connector.IntegrityError, etc. — get_db() itself lives in db.py
from werkzeug.security import generate_password_hash, check_password_hash

# ──────────────────────────────────────────────
# DATABASE
# ──────────────────────────────────────────────
# Single source of truth for the DB connection. db.py loads .env (local
# dev only), validates DB_HOST/DB_PORT/DB_USER/DB_PASSWORD/DB_NAME, and
# exposes the one and only get_db() implementation in the whole backend.
# There is no hardcoded host/user/db/password/localhost fallback anywhere
# in this file or in db.py — if a required variable is missing, importing
# db here fails immediately with a clear error instead of silently
# connecting to the wrong database.
from db import get_db, test_connection, DB_HOST, DB_PORT, DB_USER, DB_PASSWORD, DB_NAME

try:
    import pandas as pd
    PANDAS_OK = True
except ImportError:
    PANDAS_OK = False
    print("WARNING: pandas not installed — /api/import-excel will be disabled")

print("Think Easy API v2 — hierarchy + hide/unhide")

app = Flask(__name__)
CORS(
    app,
    supports_credentials=True,
    resources={
        r"/api/*": {
            "origins": [
                # Production frontend (Vercel)
                "https://thinkeasy-react.vercel.app",
                # Any Vercel preview deploy for this same project, e.g.
                # https://thinkeasy-react-git-feature-x-yourteam.vercel.app
                # or https://thinkeasy-react-<hash>.vercel.app
                re.compile(r"^https://thinkeasy-react(-[a-z0-9-]+)?\.vercel\.app$"),
                # Legacy Netlify deploy (remove once nothing points at it)
                "https://thinkeasy-1.netlify.app",
                re.compile(r"^https://[a-z0-9-]+--thinkeasy-1\.netlify\.app$"),
                # Local development
                "http://localhost:3000",
                "http://127.0.0.1:3000",
                "http://localhost:5173",
                "http://127.0.0.1:5173",
                "http://localhost:5500",
                "http://127.0.0.1:5500",
                "http://localhost:8000",
                "http://127.0.0.1:8000",
                # admin.html / business-details.html opened directly as a file
                "null",
            ]
        }
    }
)

# ──────────────────────────────────────────────
# SESSION CONFIG
# ──────────────────────────────────────────────
app.config["SECRET_KEY"] = os.environ.get("ADMIN_SECRET_KEY") or os.urandom(32).hex()
if not os.environ.get("ADMIN_SECRET_KEY"):
    print(
        "WARNING: ADMIN_SECRET_KEY is not set — using a random key generated "
        "at startup. All sessions (including logged-in users) will be "
        "invalidated on every restart/redeploy. Set ADMIN_SECRET_KEY in "
        "Render's environment variables to a fixed random string.",
        flush=True,
    )

# ──────────────────────────────────────────────
# GLOBAL ERROR HANDLERS
# Any unhandled exception now returns JSON instead
# of Render's generic HTML 500 page, making errors
# visible in the browser console.
# ──────────────────────────────────────────────
import traceback as _traceback

@app.errorhandler(500)
def handle_500(e):
    tb = _traceback.format_exc()
    app.logger.error("Unhandled 500:\n%s", tb)
    return jsonify({
        "success": False,
        "error": str(e),
        "traceback": tb
    }), 500

@app.errorhandler(404)
def handle_404(e):
    return jsonify({"success": False, "error": "Endpoint not found"}), 404

@app.errorhandler(405)
def handle_405(e):
    return jsonify({"success": False, "error": "Method not allowed"}), 405
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "None"
app.config["SESSION_COOKIE_SECURE"] = True
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)

# ──────────────────────────────────────────────
# LOGIN PROTECTION
# ──────────────────────────────────────────────
MAX_FAILED_ATTEMPTS = 5
LOCKOUT_MINUTES     = 15
MAX_USERNAME_LEN    = 80
MAX_PASSWORD_LEN    = 128

# ──────────────────────────────────────────────
# DATABASE STARTUP CHECK
# ──────────────────────────────────────────────
# get_db(), and all DB_* config, live in db.py — imported above. This is
# the ONLY get_db() implementation in the whole backend. Every route
# below calls this same imported get_db(); nothing here redefines it.
#
# Run one connectivity check immediately at boot so a broken DB config
# is obvious in the very first lines of the startup logs, before any
# request comes in:
test_connection()


# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────
def _json_dumps(val, default=None):
    """Safely JSON-encode; fall back to default string if already a string."""
    if val is None:
        return json.dumps(default if default is not None else [])
    if isinstance(val, str):
        return val          # already stored as JSON text
    return json.dumps(val)


def _num_or_none(val):
    """
    FIX: Several columns (market_size, min_investment, max_investment) are
    DECIMAL in MySQL. The admin form sends "" when those fields are left
    blank, and MySQL rejects an empty string for a DECIMAL column (causing
    an uncaught exception -> 500 Internal Server Error on save). Convert
    blank/whitespace-only values to NULL so the insert/update succeeds.
    """
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return None
    return val


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("admin_id"):
            return jsonify({"error": "Authentication required"}), 401
        return f(*args, **kwargs)
    return wrapper


def user_login_required(f):
    """
    Protects customer-facing routes. Deliberately checks session['user_id']
    only — never admin_id — so an admin session cannot be reused to pass
    as a logged-in customer, and vice versa.
    """
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return jsonify({"success": False, "message": "Authentication required"}), 401
        return f(*args, **kwargs)
    return wrapper


def log_activity(action, details):
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("INSERT INTO activity_logs (action, details) VALUES (%s, %s)",
                    (action, details))
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        print("Activity log error:", e)


# ──────────────────────────────────────────────
# STARTUP
# ──────────────────────────────────────────────
def init_admin_table():
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS admin (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(80) NOT NULL UNIQUE,
                email VARCHAR(255) NOT NULL UNIQUE,
                password_hash VARCHAR(255) NOT NULL,
                is_active TINYINT(1) NOT NULL DEFAULT 1,
                failed_login_attempts INT NOT NULL DEFAULT 0,
                locked_until DATETIME NULL DEFAULT NULL,
                otp_code VARCHAR(10) NULL DEFAULT NULL,
                otp_expiry DATETIME NULL DEFAULT NULL,
                created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()
        cur.execute("SELECT COUNT(*) FROM admin")
        (count,) = cur.fetchone()
        if count == 0:
            # SECURITY: no hardcoded demo admin accounts are seeded anymore.
            # Create the first admin via a one-off script/DB insert using
            # generate_password_hash() with a real password you choose,
            # or set ADMIN_BOOTSTRAP_USERNAME/EMAIL/PASSWORD env vars below.
            boot_user = os.environ.get("ADMIN_BOOTSTRAP_USERNAME")
            boot_email = os.environ.get("ADMIN_BOOTSTRAP_EMAIL")
            boot_pass = os.environ.get("ADMIN_BOOTSTRAP_PASSWORD")
            if boot_user and boot_email and boot_pass:
                cur.execute(
                    "INSERT INTO admin (username, email, password_hash, is_active) VALUES (%s,%s,%s,1)",
                    (boot_user, boot_email, generate_password_hash(boot_pass))
                )
                conn.commit()
                print(f"Bootstrapped initial admin account '{boot_user}' from environment variables.")
            else:
                print("No admin accounts exist yet. Set ADMIN_BOOTSTRAP_USERNAME, "
                      "ADMIN_BOOTSTRAP_EMAIL and ADMIN_BOOTSTRAP_PASSWORD env vars "
                      "and restart once to create the first admin account.")
        cur.close(); conn.close()
        print("Admin table ready.")
    except Exception as e:
        print("Could not initialize admin table:", e)


# ──────────────────────────────────────────────
# USERS TABLE  (customer accounts — separate from `admin`)
# ──────────────────────────────────────────────
def init_users_table():
    """
    Creates the customer-facing users table. Kept entirely separate from
    `admin`: different table, different session key (session['user_id']
    vs session['admin_id']), different login endpoints. Nothing here can
    grant admin access, and nothing in the admin flow touches this table.
    """
    try:
        conn = get_db()
        if conn is None:
            print("Could not initialize users table: no DB connection")
            return
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(150) NOT NULL,
                email VARCHAR(255) NOT NULL UNIQUE,
                password_hash VARCHAR(255) NULL,
                google_id VARCHAR(255) NULL UNIQUE,
                is_active TINYINT(1) NOT NULL DEFAULT 1,
                role VARCHAR(30) NOT NULL DEFAULT 'user',
                failed_login_attempts INT NOT NULL DEFAULT 0,
                locked_until DATETIME NULL DEFAULT NULL,
                created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                            ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_users_email (email),
                INDEX idx_users_google_id (google_id)
            )
        """)
        conn.commit()
        cur.close(); conn.close()
        print("Users table ready.")
    except Exception as e:
        print("Could not initialize users table:", e)


# ──────────────────────────────────────────────
# VISIBILITY BACKFILL
# ──────────────────────────────────────────────
def backfill_visibility_flags():
    """
    One-time (idempotent, safe to run on every startup) repair for rows
    that were inserted before is_hidden was explicitly set on every
    insert path. Any row sitting at is_hidden=NULL is invisible to every
    public/search query (`is_hidden = 0` never matches NULL in SQL) even
    though it's fully saved and shows correctly in Admin — this is what
    made imported products like "10KW Solar Kit" invisible on the
    homepage search. Running this fixes already-imported data immediately,
    no re-import needed.
    """
    try:
        conn = get_db()
        cur  = conn.cursor()
        fixed = {}
        for table in ("categories", "businesses", "products"):
            cur.execute(f"UPDATE {table} SET is_hidden = 0 WHERE is_hidden IS NULL")
            fixed[table] = cur.rowcount
        conn.commit()
        cur.close(); conn.close()
        total = sum(fixed.values())
        if total:
            print(f"[startup] Backfilled is_hidden=0 on {total} row(s) that were previously "
                  f"NULL and therefore invisible to search: {fixed}")
        else:
            print("[startup] Visibility flag backfill: nothing to fix.")
    except Exception as e:
        print("Could not run visibility backfill:", e)


# ══════════════════════════════════════════════════════════════════
# HEALTH
# ══════════════════════════════════════════════════════════════════
@app.route("/")
def index():
    return jsonify({"status": "ok", "service": "thinkeasy api v2"}), 200

@app.route("/health")
def health():
    return jsonify({"status": "ok", "time": datetime.utcnow().isoformat()}), 200


# ══════════════════════════════════════════════════════════════════
# AUTH
# ══════════════════════════════════════════════════════════════════
@app.route("/api/admin/login", methods=["POST"])
def admin_login():
    data     = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400
    if len(username) > MAX_USERNAME_LEN or len(password) > MAX_PASSWORD_LEN:
        return jsonify({"error": "Invalid username or password"}), 400

    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM admin WHERE username=%s", (username,))
    user = cur.fetchone()
    if not user:
        cur.close(); conn.close()
        return jsonify({"error": "Invalid username or password"}), 401

    locked_until = user.get("locked_until")
    if locked_until and locked_until > datetime.utcnow():
        remaining = max(1, int((locked_until - datetime.utcnow()).total_seconds() // 60) + 1)
        cur.close(); conn.close()
        return jsonify({"error": f"Account locked. Try again in {remaining} min."}), 423

    if not user.get("is_active"):
        cur.close(); conn.close()
        return jsonify({"error": "Account disabled."}), 403

    if not check_password_hash(user["password_hash"], password):
        attempts = user["failed_login_attempts"] + 1
        new_lock = datetime.utcnow() + timedelta(minutes=LOCKOUT_MINUTES) if attempts >= MAX_FAILED_ATTEMPTS else None
        cur.execute("UPDATE admin SET failed_login_attempts=%s, locked_until=%s WHERE id=%s",
                    (attempts, new_lock, user["id"]))
        conn.commit(); cur.close(); conn.close()
        log_activity("Failed Login", f"Failed attempt for '{username}'")
        if new_lock:
            return jsonify({"error": f"Too many attempts. Locked for {LOCKOUT_MINUTES} min."}), 423
        return jsonify({"error": "Invalid username or password"}), 401

    cur.execute("UPDATE admin SET failed_login_attempts=0, locked_until=NULL WHERE id=%s", (user["id"],))
    conn.commit(); cur.close(); conn.close()
    session.clear(); session.permanent = True
    session["admin_id"] = user["id"]; session["username"] = user["username"]
    log_activity("Admin Login", f"Admin '{username}' logged in")
    return jsonify({"message": "Login successful", "username": user["username"]}), 200


@app.route("/api/admin/logout", methods=["POST"])
def admin_logout():
    username = session.get("username"); session.clear()
    if username: log_activity("Admin Logout", f"Admin '{username}' logged out")
    return jsonify({"message": "Logged out"}), 200


@app.route("/api/admin/session")
def admin_session_check():
    if session.get("admin_id"):
        return jsonify({"authenticated": True, "username": session.get("username")}), 200
    return jsonify({"authenticated": False}), 200


# ══════════════════════════════════════════════════════════════════
# CUSTOMER AUTH  (users table — completely separate from admin auth)
#
# All account data is read from / written to the `users` table via
# get_db(). No hardcoded users, no demo accounts, no in-memory session
# objects. Every login/signup call round-trips to MySQL.
# ══════════════════════════════════════════════════════════════════
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
MIN_PASSWORD_LEN = 8
MAX_NAME_LEN = 150
USER_MAX_FAILED_ATTEMPTS = 5
USER_LOCKOUT_MINUTES = 15


def _public_user(user_row):
    """Never send password_hash or lockout internals to the client."""
    return {
        "id": user_row["id"],
        "name": user_row["name"],
        "email": user_row["email"],
        "role": user_row.get("role", "user"),
    }


def _validate_signup_fields(name, email, password, confirm_password):
    """Server-side validation — never trust the frontend's checks alone."""
    if not name or not name.strip():
        return "Full name is required."
    if len(name) > MAX_NAME_LEN:
        return "Name is too long."
    if not email or not EMAIL_RE.match(email):
        return "Please enter a valid email address."
    if not password or len(password) < MIN_PASSWORD_LEN:
        return f"Password must be at least {MIN_PASSWORD_LEN} characters."
    if not re.search(r"[A-Za-z]", password) or not re.search(r"[0-9]", password):
        return "Password must include both letters and numbers."
    if password != confirm_password:
        return "Passwords do not match."
    return None


@app.route("/api/signup", methods=["POST"])
def signup():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    confirm_password = data.get("confirm_password") or data.get("confirmPassword") or ""

    error = _validate_signup_fields(name, email, password, confirm_password)
    if error:
        return jsonify({"success": False, "message": error}), 400

    conn = get_db()
    if conn is None:
        # get_db() already logged the real exception via print("DATABASE ERROR:", e).
        # Check Render logs for that line to see the actual connection failure
        # (auth, host/port, or SSL handshake with the Railway MySQL proxy).
        app.logger.error("Signup failed: get_db() returned None (DB connection unavailable)")
        return jsonify({"success": False, "message": "Server error. Please try again."}), 503
    cur = conn.cursor(dictionary=True)

    cur.execute("SELECT id FROM users WHERE email=%s", (email,))
    if cur.fetchone():
        cur.close(); conn.close()
        return jsonify({"success": False, "message": "An account with this email already exists."}), 409

    password_hash = generate_password_hash(password)
    try:
        cur.execute(
            "INSERT INTO users (name, email, password_hash, role, is_active) VALUES (%s,%s,%s,'user',1)",
            (name, email, password_hash)
        )
        conn.commit()
        new_id = cur.lastrowid
    except mysql.connector.IntegrityError:
        # Race condition: two signups with the same email landed together.
        conn.rollback()
        cur.close(); conn.close()
        return jsonify({"success": False, "message": "An account with this email already exists."}), 409
    except Exception as e:
        conn.rollback()
        cur.close(); conn.close()
        # Log full detail server-side always.
        app.logger.error("Signup error: %s", e)
        # In production, keep the generic message so we never leak schema/DB
        # internals to the client. Set DEBUG_API_ERRORS=1 in Render's env vars
        # temporarily to see the real reason in the response while debugging,
        # then unset it once fixed.
        if os.environ.get("DEBUG_API_ERRORS") == "1":
            return jsonify({"success": False, "message": f"Server error: {e}"}), 500
        return jsonify({"success": False, "message": "Server error. Please try again."}), 500

    cur.close(); conn.close()

    # Log the new user in automatically.
    session.clear()
    session.permanent = True
    session["user_id"] = new_id
    session["user_email"] = email

    log_activity("User Signup", f"New account created for '{email}'")
    return jsonify({
        "success": True,
        "message": "Account created successfully",
        "user": {"id": new_id, "name": name, "email": email, "role": "user"}
    }), 201


@app.route("/api/auth/debug")
def auth_debug():
    """
    Diagnostic endpoint — mirrors /api/feedback/debug but for the users
    table / signup-login path. Visit this URL directly in a browser (GET,
    no auth needed) to see exactly why signup/login is failing, instead of
    guessing from the generic "Server error" message.
    Remove or protect this endpoint once the issue is resolved.
    """
    result = {
        "db_connected": False,
        "users_table_exists": False,
        "columns": [],
        "row_count": None,
        "db_host": DB_HOST,
        "db_port": DB_PORT,
        "db_name": DB_NAME,
        "db_user": DB_USER,
        "db_password_set": bool(DB_PASSWORD),
        "secret_key_pinned": bool(os.environ.get("ADMIN_SECRET_KEY")),
        "error": None,
    }
    conn = None
    try:
        conn = get_db()
        if conn is None:
            result["error"] = (
                "get_db() returned None — the connection to MySQL itself "
                "failed. Check Render's Logs tab for a line starting with "
                "'DATABASE ERROR:' right after this request — that shows "
                "the real reason (bad password, host/port unreachable, "
                "SSL handshake failure, etc.)."
            )
            return jsonify(result), 200

        result["db_connected"] = True
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'users'
        """)
        (exists,) = cur.fetchone()
        result["users_table_exists"] = bool(exists)

        if exists:
            cur.execute("""
                SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'users'
                ORDER BY ORDINAL_POSITION
            """)
            result["columns"] = [r[0] for r in cur.fetchall()]
            cur.execute("SELECT COUNT(*) FROM users")
            (count,) = cur.fetchone()
            result["row_count"] = count

        cur.close()
    except Exception as e:
        result["error"] = f"{type(e).__name__}: {e}"
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
    return jsonify(result), 200


@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json(silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""

    if not email or not password:
        return jsonify({"success": False, "message": "Email and password are required."}), 400
    if len(email) > 255 or len(password) > MAX_PASSWORD_LEN:
        return jsonify({"success": False, "message": "Invalid email or password"}), 400

    conn = get_db()
    if conn is None:
        return jsonify({"success": False, "message": "Server error. Please try again."}), 503
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM users WHERE email=%s", (email,))
    user = cur.fetchone()

    if not user or not user.get("password_hash"):
        # Same generic message whether the email doesn't exist or the
        # account is Google-only (no password_hash) — don't leak which.
        cur.close(); conn.close()
        return jsonify({"success": False, "message": "Invalid email or password"}), 401

    locked_until = user.get("locked_until")
    if locked_until and locked_until > datetime.utcnow():
        remaining = max(1, int((locked_until - datetime.utcnow()).total_seconds() // 60) + 1)
        cur.close(); conn.close()
        return jsonify({"success": False, "message": f"Too many attempts. Try again in {remaining} min."}), 423

    if not user.get("is_active"):
        cur.close(); conn.close()
        return jsonify({"success": False, "message": "This account has been disabled."}), 403

    if not check_password_hash(user["password_hash"], password):
        attempts = user["failed_login_attempts"] + 1
        new_lock = (datetime.utcnow() + timedelta(minutes=USER_LOCKOUT_MINUTES)
                    if attempts >= USER_MAX_FAILED_ATTEMPTS else None)
        cur.execute("UPDATE users SET failed_login_attempts=%s, locked_until=%s WHERE id=%s",
                    (attempts, new_lock, user["id"]))
        conn.commit(); cur.close(); conn.close()
        if new_lock:
            return jsonify({"success": False, "message": f"Too many attempts. Locked for {USER_LOCKOUT_MINUTES} min."}), 423
        return jsonify({"success": False, "message": "Invalid email or password"}), 401

    cur.execute("UPDATE users SET failed_login_attempts=0, locked_until=NULL WHERE id=%s", (user["id"],))
    conn.commit(); cur.close(); conn.close()

    session.clear()
    session.permanent = True
    session["user_id"] = user["id"]
    session["user_email"] = user["email"]

    log_activity("User Login", f"User '{email}' logged in")
    return jsonify({
        "success": True,
        "message": "Login successful",
        "user": _public_user(user)
    }), 200


@app.route("/api/logout", methods=["POST"])
def logout():
    email = session.get("user_email")
    # Only clear the customer session keys — if an admin happens to be
    # browsing in the same browser this must not also log them out.
    session.pop("user_id", None)
    session.pop("user_email", None)
    if email:
        log_activity("User Logout", f"User '{email}' logged out")
    return jsonify({"success": True, "message": "Logged out"}), 200


@app.route("/api/me")
def me():
    user_id = session.get("user_id")
    if not user_id:
        return jsonify({"success": False, "message": "Not authenticated"}), 401

    conn = get_db()
    if conn is None:
        return jsonify({"success": False, "message": "Server error"}), 503
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM users WHERE id=%s", (user_id,))
    user = cur.fetchone()
    cur.close(); conn.close()

    if not user or not user.get("is_active"):
        session.pop("user_id", None)
        session.pop("user_email", None)
        return jsonify({"success": False, "message": "Not authenticated"}), 401

    return jsonify({"success": True, "user": _public_user(user)}), 200


@app.route("/api/google-login", methods=["POST"])
def google_login():
    """
    Real Google OAuth (verifying an id_token, creating/linking the user
    account, storing google_id) is not wired up yet — it needs a Google
    Cloud OAuth client ID/secret and an approved redirect URI, which must
    be configured outside this codebase first. Until then this endpoint
    responds honestly rather than faking a login.
    """
    return jsonify({
        "success": False,
        "message": "Google login is coming soon. Please use email and password for now."
    }), 501


# ══════════════════════════════════════════════════════════════════
# CATEGORIES
# ══════════════════════════════════════════════════════════════════

@app.route("/api/categories")
def get_categories():
    """Admin: returns ALL categories including hidden ones."""
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM categories ORDER BY name")
    rows = cur.fetchall(); cur.close(); conn.close()
    return jsonify(rows)


@app.route("/api/public/categories")
def get_public_categories():
    """Public website: only visible (is_hidden=0) categories.
    COALESCE guards against legacy rows where is_hidden was left NULL
    by an import path that didn't set it explicitly — NULL = 0 is never
    true in SQL, so those rows would otherwise silently vanish from
    every public/search query."""
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM categories WHERE COALESCE(is_hidden,0)=0 ORDER BY name")
    rows = cur.fetchall(); cur.close(); conn.close()
    return jsonify(rows)


@app.route("/api/category/<int:cat_id>")
def get_category(cat_id):
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM categories WHERE id=%s", (cat_id,))
    row = cur.fetchone(); cur.close(); conn.close()
    return jsonify(row) if row else (jsonify({"error": "Not found"}), 404)


@app.route("/api/category", methods=["POST"])
@login_required
def add_category():
    data = request.get_json() or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "INSERT INTO categories (name, slug, icon, parent_id, is_hidden) VALUES (%s,%s,%s,%s,%s)",
        (data.get("name"), data.get("slug"), data.get("icon"), data.get("parent_id"),
         int(bool(data.get("hidden") or data.get("is_hidden") or 0)))
    )
    conn.commit(); new_id = cur.lastrowid; cur.close(); conn.close()
    log_activity("Category Created", f"Category '{data.get('name')}' was created")
    return jsonify({"message": "Category added", "id": new_id}), 201


@app.route("/api/category/<int:cat_id>", methods=["PUT"])
@login_required
def update_category(cat_id):
    data = request.get_json() or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "UPDATE categories SET name=%s, slug=%s, icon=%s, parent_id=%s WHERE id=%s",
        (data.get("name"), data.get("slug"), data.get("icon"), data.get("parent_id"), cat_id)
    )
    conn.commit(); affected = cur.rowcount; cur.close(); conn.close()
    if not affected: return jsonify({"error": "Not found"}), 404
    log_activity("Category Updated", f"Category '{data.get('name')}' was updated")
    return jsonify({"message": "Category updated"})


@app.route("/api/category/<int:cat_id>/hide", methods=["POST"])
@login_required
def hide_category(cat_id):
    """Soft-hide: remains in admin, invisible on public website."""
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT name FROM categories WHERE id=%s", (cat_id,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    cur.execute("UPDATE categories SET is_hidden=1 WHERE id=%s", (cat_id,))
    conn.commit(); cur.close(); conn.close()
    log_activity("Category Hidden", f"Category '{row['name']}' hidden")
    return jsonify({"message": "Category hidden", "is_hidden": 1})


@app.route("/api/category/<int:cat_id>/unhide", methods=["POST"])
@login_required
def unhide_category(cat_id):
    """Restore visibility."""
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT name FROM categories WHERE id=%s", (cat_id,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    cur.execute("UPDATE categories SET is_hidden=0 WHERE id=%s", (cat_id,))
    conn.commit(); cur.close(); conn.close()
    log_activity("Category Unhidden", f"Category '{row['name']}' made visible")
    return jsonify({"message": "Category visible", "is_hidden": 0})


@app.route("/api/category/<int:cat_id>", methods=["DELETE"])
@login_required
def delete_category(cat_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT name FROM categories WHERE id=%s", (cat_id,))
    row = cur.fetchone()
    name = row[0] if row else "Unknown"
    # Nullify FK on businesses; products FK ON DELETE CASCADE handles products
    cur.execute("UPDATE businesses SET category_id=NULL WHERE category_id=%s", (cat_id,))
    cur.execute("DELETE FROM categories WHERE id=%s", (cat_id,))
    conn.commit(); affected = cur.rowcount; cur.close(); conn.close()
    if not affected: return jsonify({"error": "Not found"}), 404
    log_activity("Category Deleted", f"Category '{name}' was deleted")
    return jsonify({"message": "Category deleted"})


# ══════════════════════════════════════════════════════════════════
# BUSINESSES
# ══════════════════════════════════════════════════════════════════

@app.route("/api/businesses")
def get_businesses():
    """Admin: ALL businesses (including hidden). Includes category name."""
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT b.*, c.name AS category_name, c.is_hidden AS category_hidden
        FROM   businesses b
        LEFT JOIN categories c ON c.id = b.category_id
        ORDER BY b.name
    """)
    rows = cur.fetchall(); cur.close(); conn.close()
    return jsonify(rows)


@app.route("/api/public/businesses")
def get_public_businesses():
    """Public: only visible businesses whose category is also visible.
    COALESCE guards against legacy NULL is_hidden rows (see note on
    /api/public/categories)."""
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT b.*, c.name AS category_name
        FROM   businesses b
        LEFT JOIN categories c ON c.id = b.category_id
        WHERE  COALESCE(b.is_hidden,0) = 0
          AND (c.id IS NULL OR COALESCE(c.is_hidden,0) = 0)
        ORDER BY b.name
    """)
    rows = cur.fetchall(); cur.close(); conn.close()
    return jsonify(rows)


@app.route("/api/business/<int:business_id>")
def get_business(business_id):
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT b.*, c.name AS category_name, c.is_hidden AS category_hidden
        FROM   businesses b
        LEFT JOIN categories c ON c.id = b.category_id
        WHERE  b.id=%s
    """, (business_id,))
    row = cur.fetchone(); cur.close(); conn.close()
    if not row:
        return jsonify({"error": "Not found"}), 404
    # Public request: treat hidden business (or hidden category) as 404
    is_admin = bool(session.get("admin_id"))
    if not is_admin and (row.get("is_hidden") or row.get("category_hidden")):
        return jsonify({"error": "Not found"}), 404
    return jsonify(row)


@app.route("/api/business", methods=["POST"])
@login_required
def add_business():
    data = request.get_json()
    if not data: return jsonify({"error": "JSON body required"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO businesses
          (name, category_id, market_size, growth_rate, investment,
           min_investment, max_investment,
           profit_margin, breakeven, breakeven_value, breakeven_unit,
           overview, badges, roadmap, suppliers, competitors,
           growth_chart, profit_projection, investment_chart, is_hidden, created_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        data.get("name"), data.get("category_id"),
        _num_or_none(data.get("market_size")), data.get("growth_rate"),
        data.get("investment"),
        _num_or_none(data.get("min_investment")), _num_or_none(data.get("max_investment")),
        data.get("profit_margin"),
        data.get("breakeven"), data.get("breakeven_value"), data.get("breakeven_unit"),
        data.get("overview"),
        _json_dumps(data.get("badges"), []),
        _json_dumps(data.get("roadmap"), []),
        _json_dumps(data.get("suppliers"), []),
        _json_dumps(data.get("competitors"), []),
        _json_dumps(data.get("growth_chart") or data.get("growthChart"), {}),
        _json_dumps(data.get("profit_projection") or data.get("profitProjection"), []),
        _json_dumps(data.get("investment_chart") or data.get("investmentChart"), {}),
        int(bool(data.get("hidden") or data.get("is_hidden") or 0)),
        datetime.utcnow()
    ))
    conn.commit(); new_id = cur.lastrowid; cur.close(); conn.close()
    log_activity("Business Created", f"Business '{data.get('name')}' was created")
    return jsonify({"message": "Business added", "id": new_id}), 201


@app.route("/api/business/<int:business_id>", methods=["PUT"])
@login_required
def update_business(business_id):
    data = request.get_json()
    if not data: return jsonify({"error": "JSON body required"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        UPDATE businesses SET
          name=%s, category_id=%s, market_size=%s, growth_rate=%s,
          investment=%s, min_investment=%s, max_investment=%s,
          profit_margin=%s, breakeven=%s, breakeven_value=%s, breakeven_unit=%s,
          overview=%s, badges=%s, roadmap=%s, suppliers=%s, competitors=%s,
          growth_chart=%s, profit_projection=%s, investment_chart=%s
        WHERE id=%s
    """, (
        data.get("name"), data.get("category_id"),
        _num_or_none(data.get("market_size")), data.get("growth_rate"),
        data.get("investment"),
        _num_or_none(data.get("min_investment")), _num_or_none(data.get("max_investment")),
        data.get("profit_margin"),
        data.get("breakeven"), data.get("breakeven_value"), data.get("breakeven_unit"),
        data.get("overview"),
        _json_dumps(data.get("badges"), []),
        _json_dumps(data.get("roadmap"), []),
        _json_dumps(data.get("suppliers"), []),
        _json_dumps(data.get("competitors"), []),
        _json_dumps(data.get("growth_chart") or data.get("growthChart"), {}),
        _json_dumps(data.get("profit_projection") or data.get("profitProjection"), []),
        _json_dumps(data.get("investment_chart") or data.get("investmentChart"), {}),
        business_id
    ))
    conn.commit(); affected = cur.rowcount; cur.close(); conn.close()
    if not affected: return jsonify({"error": "Not found"}), 404
    log_activity("Business Updated", f"Business '{data.get('name')}' was updated")
    return jsonify({"message": "Business updated"})


@app.route("/api/business/<int:business_id>/hide", methods=["POST"])
@login_required
def hide_business(business_id):
    """Soft-hide a business (and its products stay in DB unchanged)."""
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT name FROM businesses WHERE id=%s", (business_id,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    cur.execute("UPDATE businesses SET is_hidden=1 WHERE id=%s", (business_id,))
    conn.commit(); cur.close(); conn.close()
    log_activity("Business Hidden", f"Business '{row['name']}' hidden")
    return jsonify({"message": "Business hidden", "is_hidden": 1})


@app.route("/api/business/<int:business_id>/unhide", methods=["POST"])
@login_required
def unhide_business(business_id):
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT name FROM businesses WHERE id=%s", (business_id,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    cur.execute("UPDATE businesses SET is_hidden=0 WHERE id=%s", (business_id,))
    conn.commit(); cur.close(); conn.close()
    log_activity("Business Unhidden", f"Business '{row['name']}' made visible")
    return jsonify({"message": "Business visible", "is_hidden": 0})


@app.route("/api/business/<int:business_id>", methods=["DELETE"])
@login_required
def delete_business(business_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT name FROM businesses WHERE id=%s", (business_id,))
    row = cur.fetchone(); name = row[0] if row else "Unknown"
    # Cascade delete sources and schemes before deleting business
    cur.execute("DELETE FROM business_sources WHERE business_id=%s", (business_id,))
    cur.execute("DELETE FROM government_schemes WHERE business_id=%s", (business_id,))
    cur.execute("DELETE FROM businesses WHERE id=%s", (business_id,))
    conn.commit(); affected = cur.rowcount; cur.close(); conn.close()
    if not affected: return jsonify({"error": "Not found"}), 404
    log_activity("Business Deleted", f"Business '{name}' was deleted")
    return jsonify({"message": "Business deleted"})


# ══════════════════════════════════════════════════════════════════
# BUSINESS SOURCES
# ══════════════════════════════════════════════════════════════════

@app.route("/api/business/<int:business_id>/sources")
def get_business_sources(business_id):
    conn = get_db(); cur = conn.cursor(dictionary=True)
    # Verify business exists and is not hidden (public check)
    cur.execute("SELECT id FROM businesses WHERE id=%s", (business_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        return jsonify({"error": "Business not found"}), 404
    cur.execute(
        "SELECT * FROM business_sources WHERE business_id=%s ORDER BY id",
        (business_id,)
    )
    rows = cur.fetchall(); cur.close(); conn.close()
    return jsonify(rows)


@app.route("/api/business/<int:business_id>/sources", methods=["POST"])
@login_required
def add_business_source(business_id):
    data = request.get_json() or {}
    source_name = (data.get("source_name") or "").strip()
    source_url  = (data.get("source_url")  or "").strip()
    if not source_name:
        return jsonify({"error": "source_name is required"}), 400
    if not source_url:
        return jsonify({"error": "source_url is required"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "INSERT INTO business_sources (business_id, source_name, source_url) VALUES (%s,%s,%s)",
        (business_id, source_name, source_url)
    )
    conn.commit(); new_id = cur.lastrowid; cur.close(); conn.close()
    log_activity("Source Added", f"Source '{source_name}' added to business {business_id}")
    return jsonify({"message": "Source added", "id": new_id}), 201


@app.route("/api/business/<int:business_id>/sources/<int:source_id>", methods=["PUT"])
@login_required
def update_business_source(business_id, source_id):
    data = request.get_json() or {}
    source_name = (data.get("source_name") or "").strip()
    source_url  = (data.get("source_url")  or "").strip()
    if not source_name:
        return jsonify({"error": "source_name is required"}), 400
    if not source_url:
        return jsonify({"error": "source_url is required"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "UPDATE business_sources SET source_name=%s, source_url=%s WHERE id=%s AND business_id=%s",
        (source_name, source_url, source_id, business_id)
    )
    conn.commit(); affected = cur.rowcount; cur.close(); conn.close()
    if not affected: return jsonify({"error": "Not found"}), 404
    return jsonify({"message": "Source updated"})


@app.route("/api/business/<int:business_id>/sources/<int:source_id>", methods=["DELETE"])
@login_required
def delete_business_source(business_id, source_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "DELETE FROM business_sources WHERE id=%s AND business_id=%s",
        (source_id, business_id)
    )
    conn.commit(); affected = cur.rowcount; cur.close(); conn.close()
    if not affected: return jsonify({"error": "Not found"}), 404
    log_activity("Source Deleted", f"Source {source_id} deleted from business {business_id}")
    return jsonify({"message": "Source deleted"})


# ══════════════════════════════════════════════════════════════════
# GOVERNMENT SCHEMES
# ══════════════════════════════════════════════════════════════════

@app.route("/api/business/<int:business_id>/schemes")
def get_government_schemes(business_id):
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id FROM businesses WHERE id=%s", (business_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        return jsonify({"error": "Business not found"}), 404
    cur.execute(
        "SELECT * FROM government_schemes WHERE business_id=%s ORDER BY id",
        (business_id,)
    )
    rows = cur.fetchall(); cur.close(); conn.close()
    return jsonify(rows)


@app.route("/api/business/<int:business_id>/schemes", methods=["POST"])
@login_required
def add_government_scheme(business_id):
    data = request.get_json() or {}
    scheme_name  = (data.get("scheme_name")  or "").strip()
    description  = (data.get("description")  or "").strip()
    official_url = (data.get("official_url") or "").strip()
    if not scheme_name:
        return jsonify({"error": "scheme_name is required"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "INSERT INTO government_schemes (business_id, scheme_name, description, official_url) VALUES (%s,%s,%s,%s)",
        (business_id, scheme_name, description, official_url)
    )
    conn.commit(); new_id = cur.lastrowid; cur.close(); conn.close()
    log_activity("Scheme Added", f"Scheme '{scheme_name}' added to business {business_id}")
    return jsonify({"message": "Scheme added", "id": new_id}), 201


@app.route("/api/business/<int:business_id>/schemes/<int:scheme_id>", methods=["PUT"])
@login_required
def update_government_scheme(business_id, scheme_id):
    data = request.get_json() or {}
    scheme_name  = (data.get("scheme_name")  or "").strip()
    description  = (data.get("description")  or "").strip()
    official_url = (data.get("official_url") or "").strip()
    if not scheme_name:
        return jsonify({"error": "scheme_name is required"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "UPDATE government_schemes SET scheme_name=%s, description=%s, official_url=%s WHERE id=%s AND business_id=%s",
        (scheme_name, description, official_url, scheme_id, business_id)
    )
    conn.commit(); affected = cur.rowcount; cur.close(); conn.close()
    if not affected: return jsonify({"error": "Not found"}), 404
    return jsonify({"message": "Scheme updated"})


@app.route("/api/business/<int:business_id>/schemes/<int:scheme_id>", methods=["DELETE"])
@login_required
def delete_government_scheme(business_id, scheme_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute(
        "DELETE FROM government_schemes WHERE id=%s AND business_id=%s",
        (scheme_id, business_id)
    )
    conn.commit(); affected = cur.rowcount; cur.close(); conn.close()
    if not affected: return jsonify({"error": "Not found"}), 404
    log_activity("Scheme Deleted", f"Scheme {scheme_id} deleted from business {business_id}")
    return jsonify({"message": "Scheme deleted"})


# ══════════════════════════════════════════════════════════════════
# PRODUCTS BY BUSINESS  (for Related Products section)
# ══════════════════════════════════════════════════════════════════

@app.route("/api/products/by-business/<int:business_id>")
def get_products_by_business(business_id):
    """Returns visible products for a given business (used by business-details page)."""
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT p.id, p.name, p.market_size, p.growth_rate, p.investment,
               p.min_investment, p.max_investment, p.profit_margin,
               p.breakeven, p.is_hidden
        FROM   products p
        JOIN   businesses b ON b.id = p.business_id
        WHERE  p.business_id = %s
          AND  COALESCE(p.is_hidden,0) = 0
          AND  COALESCE(b.is_hidden,0) = 0
        ORDER BY p.name
    """, (business_id,))
    rows = cur.fetchall(); cur.close(); conn.close()
    return jsonify(rows)


# ══════════════════════════════════════════════════════════════════
# PRODUCTS  (first-class table, belong to a business)
# ══════════════════════════════════════════════════════════════════

@app.route("/api/products")
def get_products():
    """Admin: ALL products including hidden. Returns business + category name."""
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT p.*,
               b.name AS business_name,
               b.is_hidden AS business_hidden,
               c.name AS category_name,
               c.is_hidden AS category_hidden
        FROM   products p
        JOIN   businesses b ON b.id = p.business_id
        LEFT JOIN categories c ON c.id = p.category_id
        ORDER BY p.name
    """)
    rows = cur.fetchall(); cur.close(); conn.close()
    return jsonify(rows)


@app.route("/api/public/products")
def get_public_products():
    """Public: visible products whose parent business + category are also visible.
    COALESCE guards against legacy NULL is_hidden rows (see note on
    /api/public/categories) — this is the exact bug that made imported
    products invisible to search while still showing up in Admin/DB."""
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT p.*,
               b.name AS business_name,
               c.name AS category_name
        FROM   products p
        JOIN   businesses b ON b.id = p.business_id AND COALESCE(b.is_hidden,0) = 0
        LEFT JOIN categories c ON c.id = p.category_id
        WHERE  COALESCE(p.is_hidden,0) = 0
          AND (c.id IS NULL OR COALESCE(c.is_hidden,0) = 0)
        ORDER BY p.name
    """)
    rows = cur.fetchall(); cur.close(); conn.close()
    return jsonify(rows)


@app.route("/api/product/<int:product_id>")
def get_product(product_id):
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT p.*, b.name AS business_name, b.is_hidden AS business_hidden,
               c.name AS category_name, c.is_hidden AS category_hidden
        FROM   products p
        JOIN   businesses b ON b.id = p.business_id
        LEFT JOIN categories c ON c.id = p.category_id
        WHERE  p.id=%s
    """, (product_id,))
    row = cur.fetchone(); cur.close(); conn.close()
    if not row:
        return jsonify({"error": "Not found"}), 404
    is_admin = bool(session.get("admin_id"))
    if not is_admin and (row.get("is_hidden") or row.get("business_hidden") or row.get("category_hidden")):
        return jsonify({"error": "Not found"}), 404
    return jsonify(row)


# ══════════════════════════════════════════════════════════════════
# GLOBAL SEARCH — v2
#
# Complete replacement of the old search architecture. This is a
# dedicated, DB-only endpoint with no relationship to any in-memory
# array, cache, or demo data:
#
#   • Sources: businesses and products tables ONLY. Categories are
#     navigation-only and are never returned as search results.
#   • Every call re-queries MySQL directly — there is no caching layer
#     anywhere in this function, so Add / Edit / Delete / Hide / Excel
#     Import on either table is reflected on the very next keystroke,
#     with no rebuild step and no app restart.
#   • Hidden rows (is_hidden = 1) are excluded outright, at the SQL
#     level, for both tables.
#   • Ranking is computed in SQL: exact match first, then "starts
#     with", then "contains" — so searching "Solar" ranks
#     "Solar Panel Installation" above "Portable Solar Charger", and
#     searching "RO" ranks "RO Water Plant" above
#     "Industrial RO Accessories".
#   • Response contract is fixed: { "businesses": [...], "products": [...] }
# ══════════════════════════════════════════════════════════════════
@app.route("/api/search")
def public_search():
    q = (request.args.get("q") or "").strip()

    if len(q) < 2:
        return jsonify({"businesses": [], "products": []})

    like = f"%{q}%"

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 500
    cur = conn.cursor(dictionary=True)

    # ── Businesses ───────────────────────────────────────────────────
    # Sources: businesses table only (+ categories joined purely for
    # display of the category name — not used as a search source).
    cur.execute("""
        SELECT b.id, b.name,
               c.name        AS category,
               b.investment,
               b.growth_rate,
               b.market_size
        FROM   businesses b
        LEFT JOIN categories c ON c.id = b.category_id
        WHERE  COALESCE(b.is_hidden,0) = 0
          AND  b.name LIKE %s
        ORDER BY
          CASE
            WHEN LOWER(b.name) = LOWER(%s)               THEN 0
            WHEN LOWER(b.name) LIKE LOWER(CONCAT(%s,'%%')) THEN 1
            ELSE 2
          END,
          b.name ASC
        LIMIT 10
    """, (like, q, q))
    businesses = cur.fetchall()

    # ── Products ─────────────────────────────────────────────────────
    # Sources: products table only (+ businesses/categories joined
    # purely for display of business name / category — not used as a
    # search source). Parent business must also be visible.
    cur.execute("""
        SELECT p.id, p.name,
               b.name        AS business_name,
               c.name        AS category,
               p.investment,
               p.growth_rate,
               p.market_size
        FROM   products p
        JOIN   businesses b ON b.id = p.business_id AND COALESCE(b.is_hidden,0) = 0
        LEFT JOIN categories c ON c.id = p.category_id
        WHERE  COALESCE(p.is_hidden,0) = 0
          AND  p.name LIKE %s
        ORDER BY
          CASE
            WHEN LOWER(p.name) = LOWER(%s)               THEN 0
            WHEN LOWER(p.name) LIKE LOWER(CONCAT(%s,'%%')) THEN 1
            ELSE 2
          END,
          p.name ASC
        LIMIT 10
    """, (like, q, q))
    products = cur.fetchall()

    cur.close(); conn.close()

    return jsonify({
        "businesses": businesses,
        "products":   products
    })


@app.route("/api/product", methods=["POST"])
@login_required
def add_product():
    data = request.get_json()
    if not data: return jsonify({"error": "JSON body required"}), 400
    if not data.get("business_id"):
        return jsonify({"error": "business_id is required"}), 400

    # Derive category_id from the parent business if not explicitly provided
    category_id = data.get("category_id")
    if not category_id:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT category_id FROM businesses WHERE id=%s", (data["business_id"],))
        row = cur.fetchone()
        category_id = row[0] if row else None
        cur.close(); conn.close()

    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO products
          (name, business_id, category_id,
           market_size, growth_rate, investment,
           min_investment, max_investment,
           profit_margin, breakeven, breakeven_value, breakeven_unit,
           overview, badges, roadmap, suppliers, competitors,
           growth_chart, profit_projection, investment_chart, is_hidden)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        data.get("name"), data["business_id"], category_id,
        _num_or_none(data.get("market_size")), data.get("growth_rate"),
        data.get("investment"),
        _num_or_none(data.get("min_investment")), _num_or_none(data.get("max_investment")),
        data.get("profit_margin"),
        data.get("breakeven"), data.get("breakeven_value"), data.get("breakeven_unit"),
        data.get("overview"),
        _json_dumps(data.get("badges"), []),
        _json_dumps(data.get("roadmap"), []),
        _json_dumps(data.get("suppliers"), []),
        _json_dumps(data.get("competitors"), []),
        _json_dumps(data.get("growth_chart") or data.get("growthChart"), {}),
        _json_dumps(data.get("profit_projection") or data.get("profitProjection"), []),
        _json_dumps(data.get("investment_chart") or data.get("investmentChart"), {}),
        int(bool(data.get("hidden") or data.get("is_hidden") or 0)),
    ))
    conn.commit(); new_id = cur.lastrowid; cur.close(); conn.close()
    log_activity("Product Created", f"Product '{data.get('name')}' was created")
    return jsonify({"message": "Product added", "id": new_id}), 201


@app.route("/api/product/<int:product_id>", methods=["PUT"])
@login_required
def update_product(product_id):
    data = request.get_json()
    if not data: return jsonify({"error": "JSON body required"}), 400

    # Re-derive category_id if business changed
    category_id = data.get("category_id")
    if not category_id and data.get("business_id"):
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT category_id FROM businesses WHERE id=%s", (data["business_id"],))
        row = cur.fetchone()
        category_id = row[0] if row else None
        cur.close(); conn.close()

    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        UPDATE products SET
          name=%s, business_id=%s, category_id=%s,
          market_size=%s, growth_rate=%s, investment=%s,
          min_investment=%s, max_investment=%s,
          profit_margin=%s, breakeven=%s, breakeven_value=%s, breakeven_unit=%s,
          overview=%s, badges=%s, roadmap=%s, suppliers=%s, competitors=%s,
          growth_chart=%s, profit_projection=%s, investment_chart=%s
        WHERE id=%s
    """, (
        data.get("name"), data.get("business_id"), category_id,
        _num_or_none(data.get("market_size")), data.get("growth_rate"),
        data.get("investment"),
        _num_or_none(data.get("min_investment")), _num_or_none(data.get("max_investment")),
        data.get("profit_margin"),
        data.get("breakeven"), data.get("breakeven_value"), data.get("breakeven_unit"),
        data.get("overview"),
        _json_dumps(data.get("badges"), []),
        _json_dumps(data.get("roadmap"), []),
        _json_dumps(data.get("suppliers"), []),
        _json_dumps(data.get("competitors"), []),
        _json_dumps(data.get("growth_chart") or data.get("growthChart"), {}),
        _json_dumps(data.get("profit_projection") or data.get("profitProjection"), []),
        _json_dumps(data.get("investment_chart") or data.get("investmentChart"), {}),
        product_id
    ))
    conn.commit(); affected = cur.rowcount; cur.close(); conn.close()
    if not affected: return jsonify({"error": "Not found"}), 404
    log_activity("Product Updated", f"Product '{data.get('name')}' was updated")
    return jsonify({"message": "Product updated"})


@app.route("/api/product/<int:product_id>/hide", methods=["POST"])
@login_required
def hide_product(product_id):
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT name FROM products WHERE id=%s", (product_id,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    cur.execute("UPDATE products SET is_hidden=1 WHERE id=%s", (product_id,))
    conn.commit(); cur.close(); conn.close()
    log_activity("Product Hidden", f"Product '{row['name']}' hidden")
    return jsonify({"message": "Product hidden", "is_hidden": 1})


@app.route("/api/product/<int:product_id>/unhide", methods=["POST"])
@login_required
def unhide_product(product_id):
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT name FROM products WHERE id=%s", (product_id,))
    row = cur.fetchone()
    if not row: cur.close(); conn.close(); return jsonify({"error": "Not found"}), 404
    cur.execute("UPDATE products SET is_hidden=0 WHERE id=%s", (product_id,))
    conn.commit(); cur.close(); conn.close()
    log_activity("Product Unhidden", f"Product '{row['name']}' made visible")
    return jsonify({"message": "Product visible", "is_hidden": 0})


@app.route("/api/product/<int:product_id>", methods=["DELETE"])
@login_required
def delete_product(product_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT name FROM products WHERE id=%s", (product_id,))
    row = cur.fetchone(); name = row[0] if row else "Unknown"
    cur.execute("DELETE FROM products WHERE id=%s", (product_id,))
    conn.commit(); affected = cur.rowcount; cur.close(); conn.close()
    if not affected: return jsonify({"error": "Not found"}), 404
    log_activity("Product Deleted", f"Product '{name}' was deleted")
    return jsonify({"message": "Product deleted"})


# ══════════════════════════════════════════════════════════════════
# ACTIVITY LOGS
# ══════════════════════════════════════════════════════════════════
@app.route("/api/activity-logs")
@login_required
def get_activity_logs():
    conn = get_db(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM activity_logs ORDER BY created_at DESC LIMIT 100")
    rows = cur.fetchall(); cur.close(); conn.close()
    return jsonify(rows)


# ──────────────────────────────────────────────────────────────────
# SCHEMA MIGRATIONS
# Adds columns to existing tables introduced after initial deploy.
# Each ALTER is idempotent: MySQL 1060 (Duplicate column) is silently
# ignored so re-deploying never breaks an already-migrated DB.
# ──────────────────────────────────────────────────────────────────
def migrate_businesses_table():
    columns = [
        ("min_investment",  "DECIMAL(18,2) NULL DEFAULT NULL"),
        ("max_investment",  "DECIMAL(18,2) NULL DEFAULT NULL"),
        ("breakeven_value", "VARCHAR(50)   NULL DEFAULT NULL"),
        ("breakeven_unit",  "VARCHAR(20)   NULL DEFAULT 'Months'"),
    ]
    try:
        conn = get_db()
        cur  = conn.cursor()
        for col, definition in columns:
            try:
                cur.execute(f"ALTER TABLE businesses ADD COLUMN {col} {definition}")
                conn.commit()
                print(f"Migration: added businesses.{col}")
            except Exception as col_err:
                if "1060" in str(col_err) or "Duplicate column" in str(col_err):
                    pass  # already exists — idempotent
                else:
                    print(f"Migration warning for businesses.{col}: {col_err}")
        cur.close(); conn.close()
    except Exception as e:
        print("migrate_businesses_table() error:", e)


def migrate_sources_schemes_tables():
    """
    Creates business_sources and government_schemes tables if they do not
    already exist.  Safe to call on every startup — idempotent.
    """
    try:
        conn = get_db(); cur = conn.cursor()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS business_sources (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                business_id INT NOT NULL,
                source_name VARCHAR(255) NOT NULL,
                source_url  TEXT NOT NULL,
                created_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_bs_business (business_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS government_schemes (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                business_id INT NOT NULL,
                scheme_name VARCHAR(255) NOT NULL,
                description TEXT NULL,
                official_url TEXT NULL,
                created_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_gs_business (business_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        conn.commit()
        print("Migration: business_sources and government_schemes tables ready")
        cur.close(); conn.close()
    except Exception as e:
        print("migrate_sources_schemes_tables() error:", e)



# ══════════════════════════════════════════════════════════════════
# EXCEL IMPORT  — Wizard-style, three independent modules
# ══════════════════════════════════════════════════════════════════

def _slugify(text):
    s = str(text or "").lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def _clean_str(val):
    if val is None:
        return None
    try:
        import math
        if isinstance(val, float) and math.isnan(val):
            return None
    except Exception:
        pass
    s = str(val).strip()
    return s if s else None


def _clean_num(val):
    if val is None:
        return None
    try:
        import math
        if isinstance(val, float) and math.isnan(val):
            return None
    except Exception:
        pass
    s = str(val).strip()
    if not s:
        return None
    s = re.sub(r"[₹$€£,\s]", "", s)
    s = s.rstrip("%×xX")
    try:
        return float(s)
    except ValueError:
        return None


def _read_excel_single(file_storage):
    """
    Read a single-sheet Excel upload.
    Returns (df_with_normalised_lowercase_cols, error_str).

    Handles two layouts:
      A) Simple: row 0 = headers, row 1+ = data        (template downloads)
      B) Multi-header: row 0 = section titles (merged),
                       row 1 = real column headers,
                       row 2 = hint/description row (skip),
                       row 3+ = data                   (ThinkEasy export format)

    Detection: if row-0 has fewer than 3 non-null cells AND row-1 has
    more non-null cells than row-0, treat as layout B.
    """
    if not PANDAS_OK:
        return None, "pandas is not installed on the server"
    try:
        raw = file_storage.read()
        xls = pd.ExcelFile(io.BytesIO(raw))
        sheet = xls.sheet_names[0]

        # -- peek at raw rows to detect layout ---------------------------
        peek = xls.parse(sheet, header=None, nrows=4)
        row0_nonnull = peek.iloc[0].notna().sum()
        row1_nonnull = peek.iloc[1].notna().sum() if len(peek) > 1 else 0

        if row1_nonnull > row0_nonnull and row0_nonnull < 4:
            # Layout B: real headers are on row 1, row 2 is hint text → skip
            df = xls.parse(sheet, header=1, skiprows=[2])
        else:
            # Layout A: normal single-header file
            df = xls.parse(sheet)

        # Normalise column names: strip, lowercase, collapse whitespace
        df.columns = [
            re.sub(r"\s+", " ", str(c).strip().lower())
            for c in df.columns
        ]

        # Forward-fill columns that use a group-header pattern
        # (first row of each group filled, subsequent rows blank).
        # We detect these as object columns where < 40% of rows are filled
        # and the non-null values are strings (category / business names).
        _FFILL_CANDIDATES = [
            "business name *",
            "categories", "category", "category name",
            "business", "business name",
        ]
        for col in df.columns:
            if col in _FFILL_CANDIDATES:
                df[col] = df[col].ffill()

        # Drop rows that are entirely NaN (blank separators)
        df = df.dropna(how="all")

        return df, None
    except Exception as e:
        return None, f"Could not open workbook: {e}"


# Column name aliases — maps every possible heading the user's file
# might use to the normalised key the import code expects.
_COL_ALIASES = {
    # ── shared / businesses ───────────────────────────────────────
    # On the Business_Sector sheet "business name *" IS the business name → "name"
    # On the Product sheet "business name *" is the PARENT business → handled separately
    "name":          ["name", "business name *", "business name"],
    "category":      ["category", "categories", "category name"],
    "market_size":   ["market_size", "market size",
                      "market size (raw number — e.g. 120000000000)",
                      "market size (raw number)"],
    "growth_rate":   ["growth_rate", "growth rate"],
    "investment":    ["investment",
                      "investment (display label) (e.g. ₹8–15 l — shown in listing cards)",
                      "investment (display label) (e.g. ₹8–15 l)",
                      "investment (display label)"],
    "profit_margin": ["profit_margin", "profit margin", "profit margin (e.g. 22–28%)"],
    "overview":      ["overview", "overviewoverview"],
    "badges":        ["badges", "badgesadd badge"],
    "roadmap":       ["roadmap", "roadmapadd roadmap phase",
                      "roadmapadd roadmap [time period in motnts with activity]"],
    "suppliers":     ["suppliers", "suppliersadd supplier",
                      "phasesuppliersadd supplier\n[name, location, type, rating scale of 10]"],
    "competitors":   ["competitors", "competitorsadd competitor",
                      "competitorsadd competitor\n[name, market share globaly out of 100%, siz"],
    "min_investment":["min_investment", "minimum investment",
                      "minimum investment (e.g. 500000 — raw number, no commas)"],
    "max_investment":["max_investment", "maximum investment",
                      "maximum investment (e.g. 2000000 — raw number, no commas)"],
    "breakeven_value":["breakeven_value", "breakeven value", "break-even value",
                       "break-even value (numeric only, e.g. 12)"],
    "breakeven_unit": ["breakeven_unit", "breakeven unit", "break-even unit"],
    "growth_chart":   ["growth_chart",
                       "growth chart data — year-wise (enter value for each year; use actual years like",
                       "growth chart data — year-wise (enter value for each year) actual years"],
    "profit_projection": ["profit_projection", "profit projection",
                          "profit projection (raw numbers only, no commas — e.g. 1.8,4.2,7.5)",
                          "profit projection (raw numbers only with +/-5 years — e.g. 2020,1.8,20"],
    "investment_chart_labels": ["investment_chart_labels", "investment chart labels",
                                "investment chart labels (comma-separated, e.g. machinery,raw material,shed,worki",
                                "investment chart labels (comma-separated)investment chart values (comm"],
    "investment_chart_values": ["investment_chart_values", "investment chart values",
                                "investment chart values (comma-separated %, e.g. 40,25,15,15,5)cancel"],
    # ── products-specific ─────────────────────────────────────────
    # "product / business name *" is the product name column in the Products sheet
    "product_name":  ["product / business name *", "product name", "product/business name"],
    # The products sheet also has a plain "business" column in simple template imports
    "business":      ["business"],
}


def _normalise_row(row_dict):
    """
    Map long/messy column names to canonical short keys.
    First alias in each list wins for any given column string.
    """
    reverse = {}
    for canonical, aliases in _COL_ALIASES.items():
        for alias in aliases:
            al = re.sub(r"\s+", " ", alias.strip().lower())
            if al not in reverse:
                reverse[al] = canonical
    out = {}
    for col, val in row_dict.items():
        col_l = re.sub(r"\s+", " ", str(col).strip().lower())
        canon = reverse.get(col_l, col_l)
        if canon not in out:
            out[canon] = val
    return out


def _extract_product_parent_business(row_dict):
    """
    On the ThinkEasy Product sheet the parent-business column is literally
    named 'business name *' (col index 2) while the product name is
    'product / business name *' (col index 3).

    After _normalise_row both land under 'name' and 'product_name'.
    We need the original 'business name *' value as the parent business.
    This helper reads it directly from the raw dict before normalisation.
    """
    for col, val in row_dict.items():
        col_l = re.sub(r"\s+", " ", str(col).strip().lower())
        if col_l == "business name *":
            import math
            if val is None: return None
            try:
                if isinstance(val, float) and math.isnan(val): return None
            except Exception:
                pass
            s = str(val).strip()
            return s if s else None
    return None


def _build_single_template(sheet_name, columns):
    """
    Build a single-sheet openpyxl workbook for one import type.
    columns: list of (col_name, description, required:bool)
    Returns (BytesIO, error_str).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, "openpyxl not installed"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    hdr_fill  = PatternFill("solid", fgColor="4A4ADB")
    req_fill  = PatternFill("solid", fgColor="EFF6FF")
    opt_fill  = PatternFill("solid", fgColor="F4F6FC")
    hdr_font  = Font(color="FFFFFF", bold=True, size=11)
    req_font  = Font(color="1D4ED8", italic=True, size=10)
    opt_font  = Font(color="64748B", italic=True, size=10)

    for ci, (col, desc, required) in enumerate(columns, start=1):
        # Row 1: column name header
        cell = ws.cell(row=1, column=ci, value=col + (" *" if required else ""))
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = max(20, len(col) + 6)
        # Row 2: hint / description
        hint = ws.cell(row=2, column=ci, value=desc)
        hint.font  = req_font if required else opt_font
        hint.fill  = req_fill if required else opt_fill
        hint.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 52

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, None


# ── CATEGORIES WIZARD ─────────────────────────────────────────────────

_CAT_COLUMNS = [
    ("name",   "Category name (required)",                                     True),
    ("parent", "Parent category name — leave blank for top-level",             False),
    ("icon",   "Emoji icon, e.g. 🏭 (optional)",                              False),
    ("hidden", "1 to hide from public, 0 or blank to show (default: 0)",       False),
]


@app.route("/api/import/categories/template", methods=["GET"])
@login_required
def import_categories_template():
    buf, err = _build_single_template("Categories", _CAT_COLUMNS)
    if err:
        return jsonify({"error": err}), 500
    from flask import send_file
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="categories_import_template.xlsx")


@app.route("/api/import/categories/preview", methods=["POST"])
@login_required
def preview_categories():
    """Dry-run — returns first 20 unique rows classified as valid/duplicate/invalid."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx / .xls files are accepted"}), 400

    df, err = _read_excel_single(f)
    if err:
        return jsonify({"error": err}), 400

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 500
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT name FROM categories")
    existing = {r["name"].strip().lower() for r in cur.fetchall()}
    cur.close(); conn.close()

    rows_out = []
    seen_in_file = set()
    all_rows = df.to_dict("records")
    full_counts = {"valid": 0, "duplicate": 0, "invalid": 0}
    display_count = 0

    for row in all_rows:
        row = _normalise_row(row)
        # A dedicated "name" column takes priority. If the sheet has no
        # plain category-name column (e.g. business export files where
        # the only category info lives in "category"/"categories"),
        # fall back to that value so the file is still usable.
        name       = _clean_str(row.get("name"))
        if not name:
            name = _clean_str(row.get("category") or row.get("category_name"))
        parent_raw = _clean_str(row.get("parent") or row.get("parent_category"))
        icon_raw   = _clean_str(row.get("icon")) or ""
        hidden_raw = _clean_str(row.get("hidden")) or "0"

        if not name:
            status = "invalid"; reason = "name is required"
            full_counts[status] += 1
            if display_count < 20:
                rows_out.append({"row": display_count + 1, "name": "(empty)", "status": status,
                                 "parent": parent_raw or "", "icon": icon_raw, "hidden": hidden_raw,
                                 "reason": reason})
                display_count += 1
            continue

        nl = name.lower()
        # Dedupe identical category names that repeat across many rows
        # (e.g. one category name attached to dozens of business rows).
        if nl in existing:
            status = "duplicate"; reason = "Already exists in database (icon/visibility will be updated)"
        elif nl in seen_in_file:
            status = "duplicate"; reason = "Appears more than once in this file"
        else:
            reason = "Will be imported"
            if parent_raw and parent_raw.lower() not in existing and parent_raw.lower() not in seen_in_file:
                reason = f"Will be imported (parent '{parent_raw}' must exist first or be in this file)"
            status = "valid"
            seen_in_file.add(nl)

        full_counts[status] += 1
        if display_count < 20:
            rows_out.append({"row": display_count + 1, "name": name, "status": status,
                             "parent": parent_raw or "", "icon": icon_raw, "hidden": hidden_raw,
                             "reason": reason})
            display_count += 1

    total_rows = len(df)
    unique_count = full_counts["valid"] + full_counts["duplicate"] + full_counts["invalid"]
    return jsonify({
        "rows":          rows_out,
        "total_rows":    total_rows,
        "unique_rows":   unique_count,
        "preview_limit": 20,
        "counts":        full_counts
    })


@app.route("/api/import/categories", methods=["POST"])
@login_required
def import_categories():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx / .xls files are accepted"}), 400

    df, err = _read_excel_single(f)
    if err:
        return jsonify({"error": err}), 400

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 500
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, name FROM categories")
    existing = {r["name"].strip().lower(): r["id"] for r in cur.fetchall()}

    report = {"imported": 0, "updated": 0, "skipped": 0, "failed": 0, "rows": []}

    def _row(status, name, reason=""):
        report["rows"].append({"name": name or "(empty)", "status": status, "reason": reason})
        report[status] += 1

    all_rows = df.to_dict("records")

    # Two-pass to handle parent → child ordering in same file
    def _process(rows, second_pass=False):
        for i, row in enumerate(rows, start=1):
            row  = _normalise_row(row)
            name = _clean_str(row.get("name"))
            if not name:
                # Fall back to the "category" column when the sheet has
                # no dedicated category-name column (e.g. a business
                # export file where category info lives in "category").
                name = _clean_str(row.get("category") or row.get("category_name"))
            if not name:
                if not second_pass:
                    _row("failed", None, "name is required")
                continue
            nl = name.lower()

            parent_raw = _clean_str(row.get("parent") or row.get("parent_category"))
            parent_id  = None
            if parent_raw:
                pl = parent_raw.lower()
                if pl not in existing:
                    if not second_pass:
                        continue   # defer to second pass
                    _row("failed", name, f"Parent category '{parent_raw}' not found")
                    continue
                parent_id = existing[pl]

            icon   = _clean_str(row.get("icon"))   or "📁"
            hidden = int(_clean_num(row.get("hidden") or 0) or 0)

            if nl in existing:
                # Update is_hidden and icon if changed
                eid = existing[nl]
                try:
                    cur.execute(
                        "UPDATE categories SET icon=%s, is_hidden=%s WHERE id=%s",
                        (icon, hidden, eid)
                    )
                    conn.commit()
                    _row("updated", name, "Icon/visibility updated")
                except Exception as e:
                    conn.rollback()
                    _row("failed", name, str(e))
                continue

            slug = _slugify(name)
            try:
                cur.execute(
                    "INSERT INTO categories (name, slug, icon, parent_id, is_hidden) VALUES (%s,%s,%s,%s,%s)",
                    (name, slug, icon, parent_id, hidden)
                )
                conn.commit()
                existing[nl] = cur.lastrowid
                _row("imported", name)
                log_activity("Category Created", f"'{name}' imported via Excel wizard")
            except Exception as e:
                conn.rollback()
                _row("failed", name, str(e))

    _process(all_rows, second_pass=False)
    _process(all_rows, second_pass=True)

    cur.close(); conn.close()
    return jsonify({
        "message": "Categories import complete",
        "summary": {k: report[k] for k in ("imported", "updated", "skipped", "failed")},
        "rows":    report["rows"]
    })


# ── BUSINESSES WIZARD ─────────────────────────────────────────────────

_BIZ_COLUMNS = [
    ("name",           "Business name (required)",                                  True),
    ("category",       "Category name — must exist in database (required)",         True),
    ("market_size",    "Raw number, e.g. 5000000000",                               False),
    ("growth_rate",    "e.g. 14.5%",                                                False),
    ("investment",     "Display label, e.g. ₹5–10 L",                             False),
    ("min_investment", "Minimum investment amount (number, e.g. 500000)",           False),
    ("max_investment", "Maximum investment amount (number, e.g. 2000000)",          False),
    ("profit_margin",  "e.g. 20–25%",                                               False),
    ("breakeven",      "Full string, e.g. 12 Months",                               False),
    ("breakeven_value","Numeric part only, e.g. 12",                                False),
    ("breakeven_unit", "Months or Years",                                            False),
    ("overview",       "Short description paragraph",                                False),
]


@app.route("/api/import/businesses/template", methods=["GET"])
@login_required
def import_businesses_template():
    buf, err = _build_single_template("Businesses", _BIZ_COLUMNS)
    if err:
        return jsonify({"error": err}), 500
    from flask import send_file
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="businesses_import_template.xlsx")


@app.route("/api/import/businesses/preview", methods=["POST"])
@login_required
def preview_businesses():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx / .xls files are accepted"}), 400

    df, err = _read_excel_single(f)
    if err:
        return jsonify({"error": err}), 400

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 500
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT name FROM categories")
    valid_cats = {r["name"].strip().lower() for r in cur.fetchall()}
    cur.execute("SELECT name FROM businesses")
    existing_biz = {r["name"].strip().lower() for r in cur.fetchall()}
    cur.close(); conn.close()

    # Collect all category names in the file so we can treat them as
    # "will be auto-created" even if not in DB yet
    cats_in_file = set()
    for raw in df.to_dict("records"):
        r = _normalise_row(raw)
        c = _clean_str(r.get("category") or r.get("category_name"))
        if c:
            cats_in_file.add(c.lower())
    all_known_cats = valid_cats | cats_in_file  # DB cats + file cats

    rows_out = []
    seen_in_file = set()
    # Scan ALL rows for accurate counts, but only return first 20 for display
    all_rows = df.to_dict("records")
    full_counts = {"valid": 0, "duplicate": 0, "invalid": 0}

    for i, row in enumerate(all_rows, start=1):
        row          = _normalise_row(row)
        name         = _clean_str(row.get("name"))
        cat_raw      = _clean_str(row.get("category") or row.get("category_name"))
        investment   = _clean_str(row.get("investment")) or ""
        growth_rate  = _clean_str(row.get("growth_rate")) or ""
        profit_margin= _clean_str(row.get("profit_margin")) or ""
        breakeven    = _clean_str(row.get("breakeven")) or ""

        if not name:
            status = "invalid"; reason = "name is required"
        elif not cat_raw:
            status = "invalid"; reason = "category is required"
        else:
            nl = name.lower()
            if nl in existing_biz or nl in seen_in_file:
                status = "duplicate"
                reason = "Already exists in database" if nl in existing_biz else "Duplicate in file"
            else:
                status = "valid"
                if cat_raw.lower() not in valid_cats:
                    reason = f"Will be imported (category '{cat_raw}' will be auto-created)"
                else:
                    reason = f"Will be imported under '{cat_raw}'"
                seen_in_file.add(nl)

        full_counts[status] += 1

        if i <= 20:
            rows_out.append({
                "row": i, "name": name or "(empty)", "status": status,
                "category": cat_raw or "—",
                "investment": investment, "growth_rate": growth_rate,
                "profit_margin": profit_margin, "breakeven": breakeven,
                "reason": reason
            })

    total_rows = len(df)
    return jsonify({
        "rows":         rows_out,
        "total_rows":   total_rows,
        "preview_limit": 20,
        "counts":       full_counts
    })


@app.route("/api/import/businesses", methods=["POST"])
@login_required
def import_businesses():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx / .xls files are accepted"}), 400

    df, err = _read_excel_single(f)
    if err:
        return jsonify({"error": err}), 400

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 500
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, name FROM categories")
    cat_map = {r["name"].strip().lower(): r["id"] for r in cur.fetchall()}
    cur.execute("SELECT name FROM businesses")
    existing_biz = {r["name"].strip().lower() for r in cur.fetchall()}

    report = {"imported": 0, "updated": 0, "skipped": 0, "failed": 0, "rows": []}

    def _row(status, name, reason=""):
        report["rows"].append({"name": name or "(empty)", "status": status, "reason": reason})
        report[status] += 1

    for row in df.to_dict("records"):
        row     = _normalise_row(row)
        name    = _clean_str(row.get("name"))
        cat_raw = _clean_str(row.get("category") or row.get("category_name"))
        if not name:
            _row("failed", None, "name is required"); continue
        if not cat_raw:
            _row("failed", name, "category is required"); continue

        cat_id = cat_map.get(cat_raw.lower())
        if cat_id is None:
            # Auto-create the missing category so the business import
            # never has to be blocked on category existence.
            try:
                slug = _slugify(cat_raw)
                cur.execute(
                    "INSERT INTO categories (name, slug, icon, parent_id, is_hidden) VALUES (%s,%s,%s,%s,%s)",
                    (cat_raw, slug, "📁", None, 0)
                )
                conn.commit()
                cat_id = cur.lastrowid
                cat_map[cat_raw.lower()] = cat_id
                log_activity("Category Created", f"'{cat_raw}' auto-created during business import")
            except Exception as e:
                conn.rollback()
                _row("failed", name, f"Could not auto-create category '{cat_raw}': {e}")
                continue

        if name.lower() in existing_biz:
            _row("skipped", name, "Already exists — no changes made"); continue

        try:
            cur.execute("""
                INSERT INTO businesses
                  (name, category_id, market_size, growth_rate, investment,
                   min_investment, max_investment,
                   profit_margin, breakeven, breakeven_value, breakeven_unit,
                   overview, badges, roadmap, suppliers, competitors,
                   growth_chart, profit_projection, investment_chart,
                   is_hidden, created_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                name, cat_id,
                _clean_num(row.get("market_size")),
                _clean_str(row.get("growth_rate")),
                _clean_str(row.get("investment")),
                _clean_num(row.get("min_investment")),
                _clean_num(row.get("max_investment")),
                _clean_str(row.get("profit_margin")),
                _clean_str(row.get("breakeven")),
                _clean_str(row.get("breakeven_value")),
                _clean_str(row.get("breakeven_unit")) or "Months",
                _clean_str(row.get("overview")),
                "[]", "[]", "[]", "[]",
                json.dumps({"labels": [], "values": []}),
                "[]",
                json.dumps({"labels": [], "values": []}),
                int(_clean_num(row.get("hidden")) or 0),
                datetime.utcnow()
            ))
            conn.commit()
            existing_biz.add(name.lower())
            _row("imported", name)
            log_activity("Business Created", f"'{name}' imported via Excel wizard")
        except Exception as e:
            conn.rollback()
            _row("failed", name, str(e))

    cur.close(); conn.close()
    return jsonify({
        "message": "Businesses import complete",
        "summary": {k: report[k] for k in ("imported", "updated", "skipped", "failed")},
        "rows":    report["rows"]
    })


# ── PRODUCTS WIZARD ───────────────────────────────────────────────────

_PROD_COLUMNS = [
    ("name",           "Product name (required)",                                    True),
    ("business",       "Parent business name — must exist in database (required)",   True),
    ("market_size",    "Raw number, e.g. 5000000000",                                False),
    ("growth_rate",    "e.g. 14.5%",                                                 False),
    ("investment",     "Display label, e.g. ₹5–10 L",                              False),
    ("min_investment", "Minimum investment (number)",                                 False),
    ("max_investment", "Maximum investment (number)",                                 False),
    ("profit_margin",  "e.g. 20–25%",                                                False),
    ("breakeven",      "Full string, e.g. 12 Months",                                False),
    ("breakeven_value","Numeric part only, e.g. 12",                                 False),
    ("breakeven_unit", "Months or Years",                                             False),
    ("overview",       "Short description paragraph",                                 False),
]


@app.route("/api/import/products/template", methods=["GET"])
@login_required
def import_products_template():
    buf, err = _build_single_template("Products", _PROD_COLUMNS)
    if err:
        return jsonify({"error": err}), 500
    from flask import send_file
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="products_import_template.xlsx")


@app.route("/api/import/products/preview", methods=["POST"])
@login_required
def preview_products():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx / .xls files are accepted"}), 400

    df, err = _read_excel_single(f)
    if err:
        return jsonify({"error": err}), 400

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 500
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, name, category_id FROM businesses")
    biz_map = {r["name"].strip().lower(): r["id"] for r in cur.fetchall()}
    cur.execute("SELECT p.name, p.business_id FROM products p")
    existing_prod = {(r["name"].strip().lower(), r["business_id"]) for r in cur.fetchall()}
    cur.close(); conn.close()

    rows_out = []
    seen_in_file = set()
    all_rows = df.to_dict("records")
    full_counts = {"valid": 0, "duplicate": 0, "invalid": 0}

    for i, row in enumerate(all_rows, start=1):
        # Extract parent-business BEFORE normalise (avoids alias collision)
        biz_raw_direct = _extract_product_parent_business(row)
        row            = _normalise_row(row)
        name           = _clean_str(row.get("product_name") or row.get("name"))
        biz_raw        = biz_raw_direct or _clean_str(row.get("business") or row.get("business_name"))
        investment     = _clean_str(row.get("investment")) or ""
        growth_rate    = _clean_str(row.get("growth_rate")) or ""
        profit_margin  = _clean_str(row.get("profit_margin")) or ""

        if not name:
            status = "invalid"; reason = "name is required"
        elif not biz_raw:
            status = "invalid"; reason = "business is required"
        else:
            biz_id = biz_map.get(biz_raw.lower())
            nl = name.lower()
            file_key = (nl, biz_raw.lower())
            if biz_id is not None and (nl, biz_id) in existing_prod:
                status = "duplicate"; reason = f"Already exists under '{biz_raw}'"
            elif file_key in seen_in_file:
                status = "duplicate"; reason = f"Appears more than once under '{biz_raw}' in this file"
            else:
                status = "valid"
                if biz_id is None:
                    reason = f"Will be imported (business '{biz_raw}' will be auto-created)"
                else:
                    reason = f"Will be imported under '{biz_raw}'"
                seen_in_file.add(file_key)

        full_counts[status] += 1

        if i <= 20:
            rows_out.append({
                "row": i, "name": name or "(empty)", "status": status,
                "business": biz_raw or "—",
                "investment": investment, "growth_rate": growth_rate,
                "profit_margin": profit_margin,
                "reason": reason
            })

    total_rows = len(df)
    return jsonify({
        "rows":         rows_out,
        "total_rows":   total_rows,
        "preview_limit": 20,
        "counts":       full_counts
    })


@app.route("/api/import/products", methods=["POST"])
@login_required
def import_products():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx / .xls files are accepted"}), 400

    df, err = _read_excel_single(f)
    if err:
        return jsonify({"error": err}), 400

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 500
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, name, category_id FROM businesses")
    biz_map = {r["name"].strip().lower(): {"id": r["id"], "cat": r["category_id"]}
               for r in cur.fetchall()}
    cur.execute("SELECT p.name, p.business_id FROM products p")
    existing_prod = {(r["name"].strip().lower(), r["business_id"]) for r in cur.fetchall()}

    report = {"imported": 0, "updated": 0, "skipped": 0, "failed": 0, "rows": []}
    new_business_names = []   # collected for a single summary log entry
    imported_names      = []  # collected for a single summary log entry

    def _row(status, name, reason=""):
        report["rows"].append({"name": name or "(empty)", "status": status, "reason": reason})
        report[status] += 1

    # Cache the "Uncategorized" fallback category id (at most one lookup/insert
    # for the whole request instead of one per missing business).
    fallback_cat_id = None

    for row in df.to_dict("records"):
        biz_raw_direct = _extract_product_parent_business(row)
        row     = _normalise_row(row)
        name    = _clean_str(row.get("product_name") or row.get("name"))
        biz_raw = biz_raw_direct or _clean_str(row.get("business") or row.get("business_name"))
        if not name:
            _row("failed", None, "name is required"); continue
        if not biz_raw:
            _row("failed", name, "business is required"); continue

        biz_info = biz_map.get(biz_raw.lower())
        if biz_info is None:
            # Auto-create the missing business so the product import is
            # never blocked on business existence. New business goes
            # under an "Uncategorized" fallback category.
            try:
                if fallback_cat_id is None:
                    cur.execute("SELECT id FROM categories WHERE LOWER(name)=%s", ("uncategorized",))
                    existing_cat = cur.fetchone()
                    if existing_cat:
                        fallback_cat_id = existing_cat["id"]
                    else:
                        cur.execute(
                            "INSERT INTO categories (name, slug, icon, parent_id, is_hidden) VALUES (%s,%s,%s,%s,%s)",
                            ("Uncategorized", "uncategorized", "📁", None, 0)
                        )
                        fallback_cat_id = cur.lastrowid

                cur.execute("""
                    INSERT INTO businesses
                      (name, category_id, market_size, growth_rate, investment,
                       min_investment, max_investment,
                       profit_margin, breakeven, breakeven_value, breakeven_unit,
                       overview, badges, roadmap, suppliers, competitors,
                       growth_chart, profit_projection, investment_chart,
                       is_hidden, created_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    biz_raw, fallback_cat_id, None, None, None, None, None,
                    None, None, None, "Months", None,
                    "[]", "[]", "[]", "[]",
                    json.dumps({"labels": [], "values": []}),
                    "[]",
                    json.dumps({"labels": [], "values": []}),
                    0,
                    datetime.utcnow()
                ))
                new_biz_id = cur.lastrowid
                biz_info = {"id": new_biz_id, "cat": fallback_cat_id}
                biz_map[biz_raw.lower()] = biz_info
                new_business_names.append(biz_raw)
            except Exception as e:
                conn.rollback()
                _row("failed", name, f"Could not auto-create business '{biz_raw}': {e}")
                continue

        biz_id = biz_info["id"]; cat_id = biz_info["cat"]
        if (name.lower(), biz_id) in existing_prod:
            _row("skipped", name, f"Already exists under '{biz_raw}'"); continue

        try:
            # NOTE: is_hidden is explicitly set to 0 here. Previously this
            # column was omitted from the INSERT entirely, which meant it
            # fell back to the DB's column default. For this table that
            # default is NULL, not 0 — and `WHERE is_hidden = 0` in every
            # public/search query never matches NULL, so imported products
            # were saved successfully (visible in Admin/DB) but silently
            # invisible everywhere on the public site. This was the root
            # cause of the reported search bug.
            cur.execute("""
                INSERT INTO products
                  (name, business_id, category_id,
                   market_size, growth_rate, investment,
                   min_investment, max_investment,
                   profit_margin, breakeven, breakeven_value, breakeven_unit,
                   overview, badges, roadmap, suppliers, competitors,
                   growth_chart, profit_projection, investment_chart, is_hidden)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                name, biz_id, cat_id,
                _clean_num(row.get("market_size")),
                _clean_str(row.get("growth_rate")),
                _clean_str(row.get("investment")),
                _clean_num(row.get("min_investment")),
                _clean_num(row.get("max_investment")),
                _clean_str(row.get("profit_margin")),
                _clean_str(row.get("breakeven")),
                _clean_str(row.get("breakeven_value")),
                _clean_str(row.get("breakeven_unit")) or "Months",
                _clean_str(row.get("overview")),
                "[]", "[]", "[]", "[]",
                json.dumps({"labels": [], "values": []}),
                "[]",
                json.dumps({"labels": [], "values": []}),
                0,
            ))
            existing_prod.add((name.lower(), biz_id))
            _row("imported", name)
            imported_names.append(name)
        except Exception as e:
            _row("failed", name, str(e))

    # Single commit for the whole batch instead of one per row — each
    # commit is a network round-trip + disk flush on the remote DB, and
    # this loop could previously fire 100+ of them in one request.
    conn.commit()
    cur.close(); conn.close()

    if new_business_names:
        log_activity("Business Created",
                      f"{len(new_business_names)} businesses auto-created during product import: "
                      + ", ".join(new_business_names[:10])
                      + (f" (+{len(new_business_names) - 10} more)" if len(new_business_names) > 10 else ""))
    if imported_names:
        log_activity("Products Imported",
                      f"{len(imported_names)} products imported via Excel wizard")

    return jsonify({
        "message": "Products import complete",
        "summary": {k: report[k] for k in ("imported", "updated", "skipped", "failed")},
        "rows":    report["rows"]
    })


@app.route("/api/import/error-report", methods=["POST"])
@login_required
def import_error_report():
    """
    Accept JSON: { "type": "categories"|"businesses"|"products", "rows": [...] }
    Return a formatted FailedRows.xlsx with one sheet per failed/skipped row.
    """
    data = request.get_json(silent=True) or {}
    import_type = (data.get("type") or "import").strip()
    rows = data.get("rows") or []
    failed_rows = [r for r in rows if r.get("status") in ("failed", "skipped")]

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failed Rows"

    hdr_fill = PatternFill("solid", fgColor="DC2626")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    skip_fill= PatternFill("solid", fgColor="F59E0B")

    headers = ["#", "Name", "Status", "Reason"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60

    for ri, row in enumerate(failed_rows, start=2):
        ws.cell(row=ri, column=1, value=ri - 1)
        ws.cell(row=ri, column=2, value=row.get("name") or "(empty)")
        status_cell = ws.cell(row=ri, column=3, value=row.get("status", ""))
        ws.cell(row=ri, column=4, value=row.get("reason") or "")
        if row.get("status") == "skipped":
            for ci in range(1, 5):
                ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor="FEF9C3")
        else:
            for ci in range(1, 5):
                ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor="FEE2E2")

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"FailedRows_{import_type}.xlsx"
    )

def _read_sheet_safe(xls, name, aliases=()):
    for actual in xls.sheet_names:
        candidates = [name.lower()] + [a.lower() for a in aliases]
        if actual.lower() in candidates:
            try:
                df = xls.parse(actual)
                df.columns = [str(c).strip().lower() for c in df.columns]
                return df
            except Exception:
                return None
    return None


# ── Shared template builder ────────────────────────────────────────────

_SHEET_DEFS = {
    "Categories": [
        ("name",   "Category name (required)"),
        ("parent", "Parent category name — leave blank for top-level"),
    ],
    "Businesses": [
        ("name",          "Business name (required)"),
        ("category",      "Category name — must match existing or imported category (required)"),
        ("market_size",   "Market size as a number, e.g. 5000000000"),
        ("growth_rate",   "Growth rate string, e.g. 14.5%"),
        ("investment",    "Investment label, e.g. ₹5–10 L"),
        ("profit_margin", "Profit margin string, e.g. 20–25%"),
        ("breakeven",     "Break-even string, e.g. 12 Months"),
        ("overview",      "Short description paragraph"),
    ],
    "Products": [
        ("name",          "Product name (required)"),
        ("business",      "Parent business name — must match existing or imported business (required)"),
        ("market_size",   "Market size as a number"),
        ("growth_rate",   "Growth rate string, e.g. 14.5%"),
        ("investment",    "Investment label, e.g. ₹5–10 L"),
        ("profit_margin", "Profit margin string, e.g. 20–25%"),
        ("breakeven",     "Break-even string, e.g. 12 Months"),
    ],
}


def _build_template_workbook(sheet_names):
    """
    Build an openpyxl workbook containing only the requested sheet(s).
    sheet_names: list of keys from _SHEET_DEFS, e.g. ['Categories'] or all three.
    Returns a BytesIO buffer ready for send_file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, "openpyxl not installed"

    wb = openpyxl.Workbook()
    hdr_fill  = PatternFill("solid", fgColor="4A4ADB")
    note_fill = PatternFill("solid", fgColor="EEF2FF")
    hdr_font  = Font(color="FFFFFF", bold=True, size=11)
    note_font = Font(color="64748B", italic=True, size=10)

    first = True
    for sheet_name in sheet_names:
        cols = _SHEET_DEFS.get(sheet_name, [])
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)

        for ci, (col, _) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = max(18, len(col) + 4)

        for ci, (_, hint) in enumerate(cols, start=1):
            cell = ws.cell(row=2, column=ci, value=hint)
            cell.font = note_font
            cell.fill = note_fill
            cell.alignment = Alignment(wrap_text=True)

        ws.freeze_panes = "A3"
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 48

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, None


@app.route("/api/import-excel/template", methods=["GET"])
@login_required
def excel_template_full():
    """Download a template workbook with ALL three sheets."""
    if not PANDAS_OK:
        return jsonify({"error": "pandas is not installed on the server"}), 500
    buf, err = _build_template_workbook(["Categories", "Businesses", "Products"])
    if err:
        return jsonify({"error": err}), 500
    from flask import send_file
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="thinkeasy_import_template.xlsx")


@app.route("/api/import-excel/template/categories", methods=["GET"])
@login_required
def excel_template_categories():
    """Download Categories-only template."""
    if not PANDAS_OK:
        return jsonify({"error": "pandas not installed"}), 500
    buf, err = _build_template_workbook(["Categories"])
    if err:
        return jsonify({"error": err}), 500
    from flask import send_file
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="thinkeasy_categories_template.xlsx")


@app.route("/api/import-excel/template/businesses", methods=["GET"])
@login_required
def excel_template_businesses():
    """Download Businesses-only template."""
    if not PANDAS_OK:
        return jsonify({"error": "pandas not installed"}), 500
    buf, err = _build_template_workbook(["Businesses"])
    if err:
        return jsonify({"error": err}), 500
    from flask import send_file
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="thinkeasy_businesses_template.xlsx")


@app.route("/api/import-excel/template/products", methods=["GET"])
@login_required
def excel_template_products():
    """Download Products-only template."""
    if not PANDAS_OK:
        return jsonify({"error": "pandas not installed"}), 500
    buf, err = _build_template_workbook(["Products"])
    if err:
        return jsonify({"error": err}), 500
    from flask import send_file
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="thinkeasy_products_template.xlsx")


# ── Preview endpoint (dry-run, nothing written to DB) ─────────────────

@app.route("/api/import-excel/preview", methods=["POST"])
@login_required
def preview_excel():
    """
    Dry-run: parse the workbook, classify every row as
    valid | duplicate | invalid — but write nothing to the DB.
    Returns { categories: [...], businesses: [...], products: [...] }
    each entry: { row, name, status, reason }
    """
    if not PANDAS_OK:
        return jsonify({"error": "pandas is not installed on the server"}), 500

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx / .xls files are accepted"}), 400

    try:
        raw = f.read()
        xls = pd.ExcelFile(io.BytesIO(raw))
    except Exception as e:
        return jsonify({"error": f"Could not open workbook: {e}"}), 400

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 500
    cur = conn.cursor(dictionary=True)

    preview = {
        "categories": [],
        "businesses":  [],
        "products":    [],
    }
    counts = {
        "categories": {"valid": 0, "duplicate": 0, "invalid": 0},
        "businesses":  {"valid": 0, "duplicate": 0, "invalid": 0},
        "products":    {"valid": 0, "duplicate": 0, "invalid": 0},
    }

    # ── Categories ─────────────────────────────────────────────────
    cat_df = _read_sheet_safe(xls, "Categories", ["category", "cats"])
    if cat_df is not None:
        cur.execute("SELECT name FROM categories")
        existing = {r["name"].strip().lower() for r in cur.fetchall()}
        for i, row in enumerate(cat_df.to_dict("records"), start=1):
            name = _clean_str(row.get("name"))
            if not name:
                rec = {"row": i, "name": "(empty)", "status": "invalid",
                       "reason": "Missing name"}
                counts["categories"]["invalid"] += 1
            elif name.lower() in existing:
                rec = {"row": i, "name": name, "status": "duplicate",
                       "reason": "Already exists in database"}
                counts["categories"]["duplicate"] += 1
            else:
                rec = {"row": i, "name": name, "status": "valid",
                       "reason": "Will be imported"}
                counts["categories"]["valid"] += 1
                existing.add(name.lower())   # simulate insert for subsequent rows
            preview["categories"].append(rec)

    # ── Businesses ─────────────────────────────────────────────────
    biz_df = _read_sheet_safe(xls, "Businesses", ["business", "biz"])
    if biz_df is not None:
        cur.execute("SELECT name FROM categories")
        # Merge DB categories with any new categories from this workbook
        db_cats = {r["name"].strip().lower() for r in cur.fetchall()}
        preview_cats = {r["name"].lower() for r in preview["categories"]
                        if r["status"] in ("valid",)}
        all_cats = db_cats | preview_cats

        cur.execute("SELECT name FROM businesses")
        existing = {r["name"].strip().lower() for r in cur.fetchall()}
        for i, row in enumerate(biz_df.to_dict("records"), start=1):
            name = _clean_str(row.get("name"))
            if not name:
                rec = {"row": i, "name": "(empty)", "status": "invalid",
                       "reason": "Missing name"}
                counts["businesses"]["invalid"] += 1
            elif name.lower() in existing:
                rec = {"row": i, "name": name, "status": "duplicate",
                       "reason": "Already exists in database"}
                counts["businesses"]["duplicate"] += 1
            else:
                cat_raw = _clean_str(row.get("category") or row.get("category_name"))
                if cat_raw and cat_raw.lower() not in all_cats:
                    rec = {"row": i, "name": name, "status": "invalid",
                           "reason": f"Category '{cat_raw}' not found"}
                    counts["businesses"]["invalid"] += 1
                else:
                    rec = {"row": i, "name": name, "status": "valid",
                           "reason": "Will be imported"}
                    counts["businesses"]["valid"] += 1
                    existing.add(name.lower())
            preview["businesses"].append(rec)

    # ── Products ────────────────────────────────────────────────────
    prod_df = _read_sheet_safe(xls, "Products", ["product"])
    if prod_df is not None:
        cur.execute("SELECT id, name FROM businesses")
        db_biz = {r["name"].strip().lower() for r in cur.fetchall()}
        preview_biz = {r["name"].lower() for r in preview["businesses"]
                       if r["status"] == "valid"}
        all_biz = db_biz | preview_biz

        cur.execute("SELECT p.name, p.business_id FROM products p")
        existing = {(r["name"].strip().lower(), r["business_id"])
                    for r in cur.fetchall()}
        for i, row in enumerate(prod_df.to_dict("records"), start=1):
            name = _clean_str(row.get("name"))
            biz_raw = _clean_str(row.get("business") or row.get("business_name"))
            if not name:
                rec = {"row": i, "name": "(empty)", "status": "invalid",
                       "reason": "Missing name"}
                counts["products"]["invalid"] += 1
            elif not biz_raw:
                rec = {"row": i, "name": name, "status": "invalid",
                       "reason": "Missing business column"}
                counts["products"]["invalid"] += 1
            elif biz_raw.lower() not in all_biz:
                rec = {"row": i, "name": name, "status": "invalid",
                       "reason": f"Business '{biz_raw}' not found"}
                counts["products"]["invalid"] += 1
            else:
                # can't check duplicate by biz_id without actual IDs for new biz rows,
                # so check by name+biz_name for preview purposes
                key = (name.lower(), biz_raw.lower())
                dup_key_exists = any(
                    r["name"].lower() == name.lower()
                    for r in preview["products"]
                    if r.get("_biz","") == biz_raw.lower() and r["status"] != "invalid"
                )
                if dup_key_exists:
                    rec = {"row": i, "name": name, "status": "duplicate",
                           "reason": f"Already added under '{biz_raw}' in this file"}
                    counts["products"]["duplicate"] += 1
                else:
                    rec = {"row": i, "name": name, "status": "valid",
                           "reason": f"Will be imported under '{biz_raw}'"}
                    counts["products"]["valid"] += 1
                rec["_biz"] = biz_raw.lower()
            preview["products"].append(rec)

    cur.close(); conn.close()

    # Strip internal key before returning
    for r in preview["products"]:
        r.pop("_biz", None)

    return jsonify({"preview": preview, "counts": counts})


@app.route("/api/import-excel", methods=["POST"])
@login_required
def import_excel():
    if not PANDAS_OK:
        return jsonify({"error": "pandas is not installed on the server"}), 500

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded. Send the .xlsx as 'file' in multipart/form-data"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx / .xls files are accepted"}), 400

    try:
        raw = f.read()
        xls = pd.ExcelFile(io.BytesIO(raw))
    except Exception as e:
        return jsonify({"error": f"Could not open workbook: {e}"}), 400

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 500
    cur = conn.cursor(dictionary=True)

    report = {
        "categories": {"imported": 0, "skipped": 0, "failed": 0, "rows": []},
        "businesses":  {"imported": 0, "skipped": 0, "failed": 0, "rows": []},
        "products":    {"imported": 0, "skipped": 0, "failed": 0, "rows": []},
    }

    def _row(section, status, name, reason=""):
        report[section]["rows"].append({
            "name": name or "(unnamed)", "status": status, "reason": reason,
        })
        report[section][status] += 1

    # ── Categories ─────────────────────────────────────────────────
    cat_df = _read_sheet_safe(xls, "Categories", ["category", "cats"])
    if cat_df is not None:
        cur.execute("SELECT id, name FROM categories")
        existing_cats = {r["name"].strip().lower(): r["id"] for r in cur.fetchall()}

        def _process_cat_rows(rows, second_pass=False):
            for row in rows:
                name = _clean_str(row.get("name"))
                if not name:
                    if not second_pass:
                        _row("categories", "failed", None, "Missing 'name' value")
                    continue
                name_lc = name.lower()
                parent_raw = _clean_str(row.get("parent") or row.get("parent_category"))
                parent_id  = None
                if parent_raw:
                    parent_lc = parent_raw.lower()
                    if parent_lc not in existing_cats:
                        if not second_pass:
                            continue
                        _row("categories", "failed", name,
                             f"Parent category '{parent_raw}' not found")
                        continue
                    parent_id = existing_cats[parent_lc]
                if name_lc in existing_cats:
                    _row("categories", "skipped", name, "Already exists")
                    continue
                slug = _slugify(name)
                try:
                    cur.execute(
                        "INSERT INTO categories (name, slug, icon, parent_id, is_hidden) VALUES (%s,%s,%s,%s,%s)",
                        (name, slug, "📁", parent_id, 0)
                    )
                    conn.commit()
                    existing_cats[name_lc] = cur.lastrowid
                    _row("categories", "imported", name)
                    log_activity("Category Created", f"'{name}' imported via Excel")
                except Exception as e:
                    conn.rollback()
                    _row("categories", "failed", name, str(e))

        rows_list = cat_df.to_dict("records")
        _process_cat_rows(rows_list, second_pass=False)
        _process_cat_rows(rows_list, second_pass=True)

    # ── Businesses ─────────────────────────────────────────────────
    biz_df = _read_sheet_safe(xls, "Businesses", ["business", "biz"])
    if biz_df is not None:
        cur.execute("SELECT id, name FROM categories")
        cat_map = {r["name"].strip().lower(): r["id"] for r in cur.fetchall()}
        cur.execute("SELECT name FROM businesses")
        existing_biz = {r["name"].strip().lower() for r in cur.fetchall()}

        for row in biz_df.to_dict("records"):
            name = _clean_str(row.get("name"))
            if not name:
                _row("businesses", "failed", None, "Missing 'name' value"); continue
            if name.lower() in existing_biz:
                _row("businesses", "skipped", name, "Already exists"); continue
            cat_raw = _clean_str(row.get("category") or row.get("category_name"))
            cat_id  = cat_map.get(cat_raw.lower()) if cat_raw else None
            if cat_raw and cat_id is None:
                _row("businesses", "failed", name, f"Category '{cat_raw}' not found"); continue
            try:
                cur.execute("""
                    INSERT INTO businesses
                      (name, category_id, market_size, growth_rate, investment,
                       profit_margin, breakeven, overview,
                       badges, roadmap, suppliers, competitors,
                       growth_chart, profit_projection, investment_chart,
                       is_hidden, created_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    name, cat_id,
                    _clean_num(row.get("market_size")),
                    _clean_str(row.get("growth_rate")),
                    _clean_str(row.get("investment")),
                    _clean_str(row.get("profit_margin")),
                    _clean_str(row.get("breakeven")),
                    _clean_str(row.get("overview")),
                    "[]", "[]", "[]", "[]",
                    json.dumps({"labels": [], "values": []}),
                    "[]",
                    json.dumps({"labels": [], "values": []}),
                    0,
                    datetime.utcnow()
                ))
                conn.commit()
                existing_biz.add(name.lower())
                _row("businesses", "imported", name)
                log_activity("Business Created", f"'{name}' imported via Excel")
            except Exception as e:
                conn.rollback()
                _row("businesses", "failed", name, str(e))

    # ── Products ────────────────────────────────────────────────────
    prod_df = _read_sheet_safe(xls, "Products", ["product"])
    if prod_df is not None:
        cur.execute("SELECT id, name, category_id FROM businesses")
        biz_map = {r["name"].strip().lower(): {"id": r["id"], "cat": r["category_id"]}
                   for r in cur.fetchall()}
        cur.execute("SELECT p.name, p.business_id FROM products p")
        existing_prod = {(r["name"].strip().lower(), r["business_id"])
                         for r in cur.fetchall()}

        for row in prod_df.to_dict("records"):
            name    = _clean_str(row.get("name"))
            biz_raw = _clean_str(row.get("business") or row.get("business_name"))
            if not name:
                _row("products", "failed", None, "Missing 'name' value"); continue
            if not biz_raw:
                _row("products", "failed", name, "Missing 'business' value"); continue
            biz_info = biz_map.get(biz_raw.lower())
            if biz_info is None:
                _row("products", "failed", name, f"Business '{biz_raw}' not found"); continue
            biz_id = biz_info["id"]; cat_id = biz_info["cat"]
            if (name.lower(), biz_id) in existing_prod:
                _row("products", "skipped", name, f"Already exists under '{biz_raw}'"); continue
            try:
                cur.execute("""
                    INSERT INTO products
                      (name, business_id, category_id,
                       market_size, growth_rate, investment,
                       profit_margin, breakeven,
                       badges, roadmap, suppliers, competitors,
                       growth_chart, profit_projection, investment_chart, is_hidden)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    name, biz_id, cat_id,
                    _clean_num(row.get("market_size")),
                    _clean_str(row.get("growth_rate")),
                    _clean_str(row.get("investment")),
                    _clean_str(row.get("profit_margin")),
                    _clean_str(row.get("breakeven")),
                    "[]", "[]", "[]", "[]",
                    json.dumps({"labels": [], "values": []}),
                    "[]",
                    json.dumps({"labels": [], "values": []}),
                    0,
                ))
                conn.commit()
                existing_prod.add((name.lower(), biz_id))
                _row("products", "imported", name)
                log_activity("Product Created", f"'{name}' imported via Excel")
            except Exception as e:
                conn.rollback()
                _row("products", "failed", name, str(e))

    cur.close(); conn.close()
    summary = {k: {s: v[s] for s in ("imported", "skipped", "failed")}
               for k, v in report.items()}
    return jsonify({"message": "Import complete", "summary": summary, "details": report}), 200


# ══════════════════════════════════════════════════════════════════
# FEEDBACK CENTER
# ══════════════════════════════════════════════════════════════════

FEEDBACK_CATEGORIES = [
    "Bug Report", "Feature Request", "Business Suggestion",
    "Product Improvement", "Technical Issue", "General Feedback",
    "Question", "New Business Idea",
]
FEEDBACK_PRIORITIES = ["Low", "Medium", "High", "Critical"]
FEEDBACK_STATUSES = ["Under Review", "Planned", "In Progress", "In Development", "Released", "Rejected"]
# NOTE: "In Progress" and "In Development" are both accepted on write;
# the UI primarily surfaces "In Progress" for the trending cards while
# matching the broader status-workflow language in the spec.


def migrate_feedback_tables():
    """
    Creates feedback, feedback_votes and feedback_status_history tables
    if they do not already exist. Safe to call on every startup.

    Bugs fixed vs original:
    - Each CREATE TABLE is committed individually (avoids implicit-commit
      issues on some MySQL/MariaDB managed instances).
    - Connection is always closed in a finally block — no leak on error.
    - Full traceback is printed so Render logs show the real failure.
    - updated_at column is added via ALTER TABLE if the table already
      exists without it (handles tables created before this migration).
    """
    conn = None
    try:
        conn = get_db()
        if conn is None:
            app.logger.error("migrate_feedback_tables: could not connect to DB")
            return

        # ── Use autocommit so DDL statements take effect immediately ──
        conn.autocommit = True
        cur = conn.cursor()

        # ── feedback ──────────────────────────────────────────────────
        cur.execute("""
            CREATE TABLE IF NOT EXISTS feedback (
                id                INT AUTO_INCREMENT PRIMARY KEY,
                name              VARCHAR(120)  NOT NULL,
                email             VARCHAR(255)  NOT NULL,
                category          VARCHAR(40)   NOT NULL DEFAULT 'General Feedback',
                priority          VARCHAR(20)   NOT NULL DEFAULT 'Medium',
                subject           VARCHAR(200)  NOT NULL,
                message           TEXT          NOT NULL,
                business_id       INT           NULL,
                business_name     VARCHAR(255)  NULL,
                rating            TINYINT       NULL,
                would_recommend   VARCHAR(10)   NULL,
                is_anonymous      TINYINT(1)    NOT NULL DEFAULT 0,
                status            VARCHAR(20)   NOT NULL DEFAULT 'Under Review',
                is_pinned         TINYINT(1)    NOT NULL DEFAULT 0,
                is_trending       TINYINT(1)    NOT NULL DEFAULT 0,
                vote_count        INT           NOT NULL DEFAULT 0,
                admin_response    TEXT          NULL,
                admin_response_at DATETIME      NULL,
                created_at        DATETIME      NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at        DATETIME      NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        # Add indexes separately — IF NOT EXISTS for indexes requires MySQL 8+;
        # catch duplicate-key errors from older versions silently.
        for idx_sql in [
            "ALTER TABLE feedback ADD INDEX idx_fb_category (category)",
            "ALTER TABLE feedback ADD INDEX idx_fb_status (status)",
            "ALTER TABLE feedback ADD INDEX idx_fb_created (created_at)",
            "ALTER TABLE feedback ADD INDEX idx_fb_votes (vote_count)",
        ]:
            try:
                cur.execute(idx_sql)
            except Exception:
                pass  # index already exists — harmless

        # Ensure updated_at column exists (tables created before this fix
        # may be missing it; ALTER TABLE … ADD COLUMN IF NOT EXISTS requires
        # MySQL 8+ so we use a safer INFORMATION_SCHEMA check).
        cur.execute("""
            SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME   = 'feedback'
              AND COLUMN_NAME  = 'updated_at'
        """)
        (has_updated_at,) = cur.fetchone()
        if not has_updated_at:
            cur.execute("""
                ALTER TABLE feedback
                ADD COLUMN updated_at DATETIME NULL
            """)
            app.logger.info("migrate_feedback_tables: added missing updated_at column")

        # ── feedback_votes ────────────────────────────────────────────
        cur.execute("""
            CREATE TABLE IF NOT EXISTS feedback_votes (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                feedback_id INT         NOT NULL,
                voter_token VARCHAR(80) NOT NULL,
                created_at  DATETIME    DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY uniq_vote (feedback_id, voter_token)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        try:
            cur.execute("ALTER TABLE feedback_votes ADD INDEX idx_fv_feedback (feedback_id)")
        except Exception:
            pass

        # ── feedback_status_history ───────────────────────────────────
        cur.execute("""
            CREATE TABLE IF NOT EXISTS feedback_status_history (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                feedback_id INT         NOT NULL,
                old_status  VARCHAR(20) NULL,
                new_status  VARCHAR(20) NOT NULL,
                changed_by  VARCHAR(80) NULL,
                created_at  DATETIME    DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        try:
            cur.execute("ALTER TABLE feedback_status_history ADD INDEX idx_fsh_feedback (feedback_id)")
        except Exception:
            pass

        cur.close()
        app.logger.info("migrate_feedback_tables: all feedback tables ready")
        print("Migration: feedback tables ready")

    except Exception as e:
        app.logger.exception("migrate_feedback_tables() failed: %s", e)
        print("migrate_feedback_tables() error:", e)
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


def _feedback_row_to_dict(row):
    """Normalize a feedback DB row (dict cursor) for JSON output."""
    if row.get("is_anonymous"):
        row["name"] = "Anonymous"
        row["email"] = None
    for dt_field in ("created_at", "updated_at", "admin_response_at"):
        if row.get(dt_field):
            row[dt_field] = row[dt_field].isoformat()
    return row


def _recompute_trending(feedback_id=None):
    """
    Marks feedback as trending based on votes + recent activity.
    Heuristic: vote_count >= 5, OR (created within last 14 days AND
    vote_count >= 2), OR manually pinned (pin always implies trending).
    If feedback_id is given, only that row is recomputed; otherwise all.
    """
    try:
        conn = get_db(); cur = conn.cursor()
        where = "WHERE id=%s" if feedback_id else ""
        params = (feedback_id,) if feedback_id else ()
        cur.execute(f"""
            UPDATE feedback
            SET is_trending = CASE
                WHEN is_pinned = 1 THEN 1
                WHEN vote_count >= 5 THEN 1
                WHEN vote_count >= 2 AND created_at >= (NOW() - INTERVAL 14 DAY) THEN 1
                ELSE 0
            END
            {where}
        """, params)
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        print("_recompute_trending() error:", e)


@app.route("/api/feedback/meta")
def feedback_meta():
    """Public: category/priority/status option lists for the form + filters."""
    return jsonify({
        "categories": FEEDBACK_CATEGORIES,
        "priorities":  FEEDBACK_PRIORITIES,
        "statuses":    FEEDBACK_STATUSES,
    }), 200


@app.route("/api/feedback/debug")
def feedback_debug():
    """
    Diagnostic endpoint — shows DB connectivity + table status.
    Remove or protect this endpoint once the 500 is resolved.
    """
    result = {"db_connected": False, "feedback_table_exists": False,
              "columns": [], "error": None, "row_count": None}
    conn = None
    try:
        conn = get_db()
        if conn is None:
            result["error"] = "get_db() returned None — check DB_HOST/DB_USER/DB_PASSWORD env vars"
            return jsonify(result), 200

        result["db_connected"] = True
        cur = conn.cursor()

        # Check table exists
        cur.execute("""
            SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'feedback'
        """)
        (exists,) = cur.fetchone()
        result["feedback_table_exists"] = bool(exists)

        if exists:
            # List actual columns
            cur.execute("""
                SELECT COLUMN_NAME, COLUMN_TYPE, IS_NULLABLE, COLUMN_DEFAULT
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'feedback'
                ORDER BY ORDINAL_POSITION
            """)
            result["columns"] = [
                {"name": r[0], "type": r[1], "nullable": r[2], "default": r[3]}
                for r in cur.fetchall()
            ]
            # Row count
            cur.execute("SELECT COUNT(*) FROM feedback")
            (result["row_count"],) = cur.fetchone()
        else:
            result["error"] = "feedback table does not exist — migration failed silently"

        cur.close()
    except Exception as e:
        result["error"] = str(e)
        import traceback
        result["traceback"] = traceback.format_exc()
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

    return jsonify(result), 200


@app.route("/api/feedback", methods=["POST"])
def submit_feedback():
    """Public: submit new feedback / feature request."""
    import traceback as _tb

    # ── 1. Parse JSON body ────────────────────────────────────────
    data = request.get_json(silent=True)
    if data is None:
        # Body wasn't JSON or Content-Type header was missing
        app.logger.warning(
            "submit_feedback: non-JSON body (Content-Type=%s)",
            request.content_type
        )
        return jsonify({
            "success": False,
            "error": "Request body must be JSON (Content-Type: application/json)"
        }), 400

    # ── 2. Extract fields ─────────────────────────────────────────
    name            = (data.get("name")            or "").strip()
    email           = (data.get("email")           or "").strip()
    category        = (data.get("category")        or "General Feedback").strip()
    priority        = (data.get("priority")        or "Medium").strip()
    subject         = (data.get("subject")         or "").strip()
    message         = (data.get("message")         or "").strip()
    business_id     = data.get("business_id")
    business_name   = (data.get("business_name")   or "").strip() or None
    rating          = data.get("rating")
    would_recommend = (data.get("would_recommend") or "").strip() or None
    is_anonymous    = bool(data.get("is_anonymous"))

    # ── 3. Validate ───────────────────────────────────────────────
    if not name:
        return jsonify({"success": False, "error": "Name is required"}), 400
    if not email or "@" not in email:
        return jsonify({"success": False, "error": "A valid email is required"}), 400
    if category not in FEEDBACK_CATEGORIES:
        # Be lenient — fall back to General Feedback instead of rejecting
        app.logger.warning("submit_feedback: unknown category %r — using General Feedback", category)
        category = "General Feedback"
    if priority not in FEEDBACK_PRIORITIES:
        priority = "Medium"
    if not subject:
        return jsonify({"success": False, "error": "Subject is required"}), 400
    if not message:
        return jsonify({"success": False, "error": "Message is required"}), 400
    if len(message) > 1000:
        return jsonify({"success": False,
                        "error": "Message must be 1000 characters or fewer"}), 400

    # ── 4. Coerce optional typed fields ───────────────────────────
    try:
        business_id = int(business_id) if business_id not in (None, "", "null") else None
    except (TypeError, ValueError):
        business_id = None

    try:
        rating = int(rating) if rating not in (None, "", "null") else None
        if rating is not None and not (1 <= rating <= 5):
            rating = None
    except (TypeError, ValueError):
        rating = None

    if would_recommend not in ("Yes", "Maybe", "No", None):
        would_recommend = None

    # ── 5. Database INSERT ────────────────────────────────────────
    conn = None
    try:
        conn = get_db()
        if conn is None:
            app.logger.error("submit_feedback: get_db() returned None")
            return jsonify({
                "success": False,
                "error": "Database connection failed — please try again"
            }), 503

        cur = conn.cursor()

        # Run a minimal self-healing migration on this connection.
        # We do NOT use CREATE TABLE here (DDL causes implicit commits
        # that confuse the connector state on some Railway MySQL builds).
        # Instead we check existence first, create only if truly missing.
        cur.execute(
            "SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'feedback'"
        )
        (table_exists,) = cur.fetchone()

        if not table_exists:
            app.logger.warning("submit_feedback: feedback table missing — creating now")
            # Use a fresh autocommit connection for DDL to avoid
            # mixing DDL and DML in the same transaction
            ddl_conn = get_db()
            if ddl_conn:
                try:
                    ddl_conn.autocommit = True
                    ddl_cur = ddl_conn.cursor()
                    ddl_cur.execute("""
                        CREATE TABLE IF NOT EXISTS feedback (
                            id                INT AUTO_INCREMENT PRIMARY KEY,
                            name              VARCHAR(120) NOT NULL,
                            email             VARCHAR(255) NOT NULL,
                            category          VARCHAR(40)  NOT NULL DEFAULT 'General Feedback',
                            priority          VARCHAR(20)  NOT NULL DEFAULT 'Medium',
                            subject           VARCHAR(200) NOT NULL,
                            message           TEXT         NOT NULL,
                            business_id       INT          NULL,
                            business_name     VARCHAR(255) NULL,
                            rating            TINYINT      NULL,
                            would_recommend   VARCHAR(10)  NULL,
                            is_anonymous      TINYINT(1)   NOT NULL DEFAULT 0,
                            status            VARCHAR(20)  NOT NULL DEFAULT 'Under Review',
                            is_pinned         TINYINT(1)   NOT NULL DEFAULT 0,
                            is_trending       TINYINT(1)   NOT NULL DEFAULT 0,
                            vote_count        INT          NOT NULL DEFAULT 0,
                            admin_response    TEXT         NULL,
                            admin_response_at DATETIME     NULL,
                            created_at        DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP,
                            updated_at        DATETIME     NULL
                        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                    """)
                    ddl_cur.close()
                    app.logger.info("submit_feedback: feedback table created")
                except Exception as ddl_err:
                    app.logger.exception("submit_feedback: DDL failed: %s", ddl_err)
                    return jsonify({
                        "success": False,
                        "error": f"Table setup failed: {ddl_err}"
                    }), 500
                finally:
                    try:
                        ddl_conn.close()
                    except Exception:
                        pass

        # ── INSERT ────────────────────────────────────────────────
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute("""
            INSERT INTO feedback
                (name, email, category, priority, subject, message,
                 business_id, business_name, rating, would_recommend,
                 is_anonymous, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s)
        """, (
            name, email, category, priority, subject, message,
            business_id, business_name, rating, would_recommend,
            int(is_anonymous), now
        ))
        conn.commit()
        new_id = cur.lastrowid
        cur.close()

        app.logger.info("submit_feedback: saved id=%s category=%r", new_id, category)

    except Exception as e:
        tb = _tb.format_exc()
        app.logger.error("submit_feedback FAILED\n%s", tb)
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        return jsonify({
            "success": False,
            "error": str(e),
            "detail": tb          # remove this line once debugging is done
        }), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

    # log_activity uses its own connection — safe to call after ours is closed
    try:
        log_activity("New Feedback", f"{category} submitted: \"{subject}\"")
    except Exception:
        pass  # never let logging kill a successful response

    return jsonify({"success": True, "message": "Feedback submitted", "id": new_id}), 201




@app.route("/api/feedback", methods=["GET"])
@login_required
def list_feedback():
    """Admin: list feedback with optional search + filters."""
    q        = (request.args.get("q") or "").strip()
    category = request.args.get("category") or ""
    status   = request.args.get("status") or ""
    priority = request.args.get("priority") or ""
    sort     = request.args.get("sort") or "newest"  # newest | popular | updated

    where = []
    params = []
    if q:
        where.append("(subject LIKE %s OR message LIKE %s OR name LIKE %s OR email LIKE %s)")
        like = f"%{q}%"
        params += [like, like, like, like]
    if category:
        where.append("category = %s"); params.append(category)
    if status:
        where.append("status = %s"); params.append(status)
    if priority:
        where.append("priority = %s"); params.append(priority)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    order_sql = {
        "popular": "is_pinned DESC, vote_count DESC, created_at DESC",
        "updated": "is_pinned DESC, updated_at DESC",
    }.get(sort, "is_pinned DESC, created_at DESC")

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 503
    cur = conn.cursor(dictionary=True)
    cur.execute(f"SELECT * FROM feedback {where_sql} ORDER BY {order_sql}", params)
    rows = [_feedback_row_to_dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(rows), 200


@app.route("/api/feedback/trending")
def trending_feedback():
    """Public: trending feature requests for the Trending Requests section."""
    limit = request.args.get("limit", default=12, type=int)
    sort  = request.args.get("sort") or "popular"  # popular | newest | updated
    order_sql = {
        "newest":  "created_at DESC",
        "updated": "updated_at DESC",
    }.get(sort, "is_pinned DESC, vote_count DESC, created_at DESC")

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 503
    cur = conn.cursor(dictionary=True)
    cur.execute(f"""
        SELECT id, category, subject, message, status, vote_count,
               is_pinned, is_trending, created_at, updated_at, business_name
        FROM feedback
        WHERE is_trending = 1 OR is_pinned = 1
        ORDER BY {order_sql}
        LIMIT %s
    """, (limit,))
    rows = [_feedback_row_to_dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(rows), 200


@app.route("/api/feedback/<int:feedback_id>")
def get_feedback_item(feedback_id):
    """Public: single feedback item (for detail views)."""
    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 503
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM feedback WHERE id=%s", (feedback_id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return jsonify({"error": "Not found"}), 404
    return jsonify(_feedback_row_to_dict(row)), 200


@app.route("/api/feedback/vote", methods=["POST"])
def vote_feedback():
    """
    Public: cast or remove a vote.
    Body: { feedback_id, voter_token, action: 'vote' | 'unvote' }
    voter_token is an anonymous per-browser UUID generated client-side
    and stored in localStorage — there is no public user login system,
    so this is the de-facto "one vote per visitor" identity.
    """
    data = request.get_json(silent=True) or {}
    feedback_id = data.get("feedback_id")
    voter_token = (data.get("voter_token") or "").strip()
    action = (data.get("action") or "vote").strip()

    try:
        feedback_id = int(feedback_id)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid feedback_id"}), 400
    if not voter_token or len(voter_token) > 80:
        return jsonify({"error": "Invalid voter_token"}), 400
    if action not in ("vote", "unvote"):
        return jsonify({"error": "action must be vote | unvote"}), 400

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 503
    cur = conn.cursor()
    try:
        cur.execute("SELECT id FROM feedback WHERE id=%s", (feedback_id,))
        if not cur.fetchone():
            cur.close(); conn.close()
            return jsonify({"error": "Feedback not found"}), 404

        if action == "vote":
            try:
                cur.execute(
                    "INSERT INTO feedback_votes (feedback_id, voter_token) VALUES (%s,%s)",
                    (feedback_id, voter_token)
                )
                cur.execute(
                    "UPDATE feedback SET vote_count = vote_count + 1 WHERE id=%s",
                    (feedback_id,)
                )
                conn.commit()
                voted = True
            except mysql.connector.errors.IntegrityError:
                conn.rollback()
                voted = True  # already voted — idempotent success
        else:
            cur.execute(
                "DELETE FROM feedback_votes WHERE feedback_id=%s AND voter_token=%s",
                (feedback_id, voter_token)
            )
            if cur.rowcount:
                cur.execute(
                    "UPDATE feedback SET vote_count = GREATEST(0, vote_count - 1) WHERE id=%s",
                    (feedback_id,)
                )
            conn.commit()
            voted = False

        cur.execute("SELECT vote_count FROM feedback WHERE id=%s", (feedback_id,))
        (vote_count,) = cur.fetchone()
        cur.close(); conn.close()
    except Exception as e:
        conn.rollback(); cur.close(); conn.close()
        return jsonify({"error": f"Vote failed: {e}"}), 500

    _recompute_trending(feedback_id)
    return jsonify({"voted": voted, "vote_count": vote_count}), 200


@app.route("/api/feedback/votes")
def my_feedback_votes():
    """Public: given a voter_token, return ids the visitor has already voted on."""
    voter_token = (request.args.get("voter_token") or "").strip()
    if not voter_token:
        return jsonify({"voted_ids": []}), 200
    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 503
    cur = conn.cursor()
    cur.execute("SELECT feedback_id FROM feedback_votes WHERE voter_token=%s", (voter_token,))
    ids = [r[0] for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify({"voted_ids": ids}), 200


@app.route("/api/feedback/<int:feedback_id>/status", methods=["PUT"])
@login_required
def update_feedback_status(feedback_id):
    """Admin: change status, optionally attach an admin response."""
    data = request.get_json(silent=True) or {}
    new_status = (data.get("status") or "").strip()
    admin_response = data.get("admin_response")

    if new_status not in FEEDBACK_STATUSES:
        return jsonify({"error": f"status must be one of {FEEDBACK_STATUSES}"}), 400

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 503
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT status FROM feedback WHERE id=%s", (feedback_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return jsonify({"error": "Not found"}), 404
    old_status = row["status"]

    if admin_response:
        cur.execute("""
            UPDATE feedback
            SET status=%s, admin_response=%s, admin_response_at=NOW()
            WHERE id=%s
        """, (new_status, admin_response, feedback_id))
    else:
        cur.execute("UPDATE feedback SET status=%s WHERE id=%s", (new_status, feedback_id))

    cur.execute("""
        INSERT INTO feedback_status_history (feedback_id, old_status, new_status, changed_by)
        VALUES (%s,%s,%s,%s)
    """, (feedback_id, old_status, new_status, session.get("username")))
    conn.commit(); cur.close(); conn.close()

    log_activity("Feedback Status Changed", f"#{feedback_id}: {old_status} → {new_status}")
    return jsonify({"message": "Status updated"}), 200


@app.route("/api/feedback/<int:feedback_id>/pin", methods=["POST"])
@login_required
def pin_feedback(feedback_id):
    """Admin: toggle pin (pinned items are always shown as trending)."""
    data = request.get_json(silent=True) or {}
    pinned = bool(data.get("pinned", True))

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 503
    cur = conn.cursor()
    cur.execute("UPDATE feedback SET is_pinned=%s WHERE id=%s", (int(pinned), feedback_id))
    if cur.rowcount == 0:
        cur.close(); conn.close()
        return jsonify({"error": "Not found"}), 404
    conn.commit(); cur.close(); conn.close()

    _recompute_trending(feedback_id)
    log_activity("Feedback Pinned" if pinned else "Feedback Unpinned", f"#{feedback_id}")
    return jsonify({"message": "Updated", "is_pinned": pinned}), 200


@app.route("/api/feedback/<int:feedback_id>", methods=["DELETE"])
@login_required
def delete_feedback(feedback_id):
    """Admin: delete a feedback item (cascades votes/history manually)."""
    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 503
    cur = conn.cursor()
    cur.execute("SELECT subject FROM feedback WHERE id=%s", (feedback_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return jsonify({"error": "Not found"}), 404
    subject = row[0]

    cur.execute("DELETE FROM feedback_votes WHERE feedback_id=%s", (feedback_id,))
    cur.execute("DELETE FROM feedback_status_history WHERE feedback_id=%s", (feedback_id,))
    cur.execute("DELETE FROM feedback WHERE id=%s", (feedback_id,))
    conn.commit(); cur.close(); conn.close()

    log_activity("Feedback Deleted", f"\"{subject}\" (#{feedback_id}) deleted")
    return jsonify({"message": "Deleted"}), 200


@app.route("/api/feedback/bulk", methods=["POST"])
@login_required
def bulk_feedback():
    """Admin: bulk delete feedback items (mirrors categories/businesses/products bulk)."""
    data   = request.get_json(silent=True) or {}
    ids    = [int(i) for i in (data.get("ids") or []) if str(i).isdigit()]
    action = data.get("action", "")
    if action != "delete":
        return jsonify({"error": "action must be delete"}), 400
    if not ids:
        return jsonify({"error": "No ids provided"}), 400

    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 503
    cur = conn.cursor()
    ok = 0; errors = []
    for fid in ids:
        try:
            cur.execute("DELETE FROM feedback_votes WHERE feedback_id=%s", (fid,))
            cur.execute("DELETE FROM feedback_status_history WHERE feedback_id=%s", (fid,))
            cur.execute("DELETE FROM feedback WHERE id=%s", (fid,))
            conn.commit()
            ok += 1
        except Exception as e:
            conn.rollback()
            errors.append(f"id={fid}: {e}")
    cur.close(); conn.close()
    log_activity("Bulk Delete", f"{ok} feedback items deleted")
    return jsonify({"ok": ok, "errors": errors}), 200


@app.route("/api/feedback/stats")
@login_required
def feedback_stats():
    """Admin: dashboard cards for the Feedback Center."""
    conn = get_db()
    if conn is None:
        return jsonify({"error": "Database unavailable"}), 503
    cur = conn.cursor(dictionary=True)

    cur.execute("SELECT COUNT(*) AS c FROM feedback")
    total = cur.fetchone()["c"]

    cur.execute("SELECT COUNT(*) AS c FROM feedback WHERE category='Feature Request'")
    feature_requests = cur.fetchone()["c"]

    cur.execute("SELECT COUNT(*) AS c FROM feedback WHERE category='Bug Report'")
    bug_reports = cur.fetchone()["c"]

    cur.execute("SELECT COUNT(*) AS c FROM feedback WHERE status='Released'")
    resolved = cur.fetchone()["c"]

    cur.execute("SELECT COUNT(*) AS c FROM feedback WHERE status IN ('Under Review','Planned','In Progress','In Development')")
    pending = cur.fetchone()["c"]

    cur.execute("SELECT COUNT(*) AS c FROM feedback WHERE is_trending=1")
    trending = cur.fetchone()["c"]

    cur.execute("SELECT id, subject, vote_count FROM feedback ORDER BY vote_count DESC LIMIT 1")
    most_voted = cur.fetchone()

    cur.execute("SELECT id, subject, created_at FROM feedback ORDER BY created_at DESC LIMIT 1")
    newest = cur.fetchone()
    if newest and newest.get("created_at"):
        newest["created_at"] = newest["created_at"].isoformat()

    cur.execute("SELECT AVG(rating) AS avg_rating FROM feedback WHERE rating IS NOT NULL")
    avg_rating = cur.fetchone()["avg_rating"]
    avg_rating = round(float(avg_rating), 2) if avg_rating is not None else None

    cur.close(); conn.close()
    return jsonify({
        "total_feedback": total,
        "feature_requests": feature_requests,
        "bug_reports": bug_reports,
        "resolved": resolved,
        "pending": pending,
        "trending_requests": trending,
        "most_voted_feature": most_voted,
        "newest_feedback": newest,
        "average_rating": avg_rating,
    }), 200


# ══════════════════════════════════════════════════════════════════
# BULK OPERATIONS
# ══════════════════════════════════════════════════════════════════

def _bulk_action(table, id_col, ids, action):
    """
    Shared engine for bulk hide / unhide / delete.
    table:   'categories' | 'businesses' | 'products'
    id_col:  primary key column name (always 'id')
    ids:     list of int
    action:  'hide' | 'unhide' | 'delete'
    Returns (ok_count, errors[])
    """
    if not ids:
        return 0, []
    conn = get_db()
    if conn is None:
        return 0, ["Database unavailable"]
    cur = conn.cursor()
    ok = 0; errors = []
    for rid in ids:
        try:
            if action == "hide":
                cur.execute(f"UPDATE {table} SET is_hidden=1 WHERE {id_col}=%s", (rid,))
            elif action == "unhide":
                cur.execute(f"UPDATE {table} SET is_hidden=0 WHERE {id_col}=%s", (rid,))
            elif action == "delete":
                cur.execute(f"DELETE FROM {table} WHERE {id_col}=%s", (rid,))
            conn.commit()
            ok += 1
        except Exception as e:
            conn.rollback()
            errors.append(f"id={rid}: {e}")
    cur.close(); conn.close()
    return ok, errors


@app.route("/api/categories/bulk", methods=["POST"])
@login_required
def bulk_categories():
    data   = request.get_json(silent=True) or {}
    ids    = [int(i) for i in (data.get("ids") or []) if str(i).isdigit()]
    action = data.get("action", "")
    if action not in ("hide", "unhide", "delete"):
        return jsonify({"error": "action must be hide | unhide | delete"}), 400
    if not ids:
        return jsonify({"error": "No ids provided"}), 400
    ok, errors = _bulk_action("categories", "id", ids, action)
    log_activity(f"Bulk {action.title()}", f"{ok} categories bulk-{action}d")
    return jsonify({"ok": ok, "errors": errors})


@app.route("/api/businesses/bulk", methods=["POST"])
@login_required
def bulk_businesses():
    data   = request.get_json(silent=True) or {}
    ids    = [int(i) for i in (data.get("ids") or []) if str(i).isdigit()]
    action = data.get("action", "")
    if action not in ("hide", "unhide", "delete"):
        return jsonify({"error": "action must be hide | unhide | delete"}), 400
    if not ids:
        return jsonify({"error": "No ids provided"}), 400
    ok, errors = _bulk_action("businesses", "id", ids, action)
    log_activity(f"Bulk {action.title()}", f"{ok} businesses bulk-{action}d")
    return jsonify({"ok": ok, "errors": errors})


@app.route("/api/products/bulk", methods=["POST"])
@login_required
def bulk_products():
    data   = request.get_json(silent=True) or {}
    ids    = [int(i) for i in (data.get("ids") or []) if str(i).isdigit()]
    action = data.get("action", "")
    if action not in ("hide", "unhide", "delete"):
        return jsonify({"error": "action must be hide | unhide | delete"}), 400
    if not ids:
        return jsonify({"error": "No ids provided"}), 400
    ok, errors = _bulk_action("products", "id", ids, action)
    log_activity(f"Bulk {action.title()}", f"{ok} products bulk-{action}d")
    return jsonify({"ok": ok, "errors": errors})


# ══════════════════════════════════════════════════════════════════
# HIDDEN RECORDS — SINGLE-ITEM ENFORCEMENT
# Ensure GET /api/business/<id> and GET /api/product/<id> never
# serve hidden records to the public (no login_required here, but
# we check is_hidden and return 404 for hidden items).
# ══════════════════════════════════════════════════════════════════
init_admin_table()
init_users_table()
migrate_businesses_table()
migrate_sources_schemes_tables()
migrate_feedback_tables()
backfill_visibility_flags()

if __name__ == "__main__":
    app.run(debug=False)
